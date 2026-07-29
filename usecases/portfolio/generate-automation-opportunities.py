#!/usr/bin/env python3
"""
Generate TCS Agentic AI — Automation Use Case Summary
Outputs: Excel (.xlsx) + HTML for hackathon demo presentation
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT_DIR = os.path.dirname(os.path.abspath(__file__))

# ═══════════════════════════════════════════════════════════════════
# DATA
# ═══════════════════════════════════════════════════════════════════

TITLE = "TCS Agentic AI — Automation Opportunities for OpenShift Operations"
SUBTITLE = "End-to-End AI-Native Automation Across ITIL Ticket Categories"

# Each row: (ticket_type, automation_possibilities, industry_pct, tcs_pct, solution_levers, key_demo_moment)
DATA = [
    (
        "Incidents",
        [
            "Pod crash detection & auto-restart",
            "Node NotReady auto-cordon & drain",
            "OOM / CPU saturation auto-scaling",
            "Certificate expiry auto-renewal alerts",
            "Storage pressure detection & PVC cleanup",
            "Network policy conflict resolution",
            "Cross-cluster incident correlation",
        ],
        "60-65%",
        "85-90%",
        [
            "AI Chat — natural language root cause analysis with fix proposals",
            "16 Dashboard Widgets — real-time health, pods-at-risk, active alerts",
            "Emergency Actions — cordon, drain, force-delete with confirmation",
            "MCP Tools — 44 tools (6 cluster health, 5 monitoring, 6 emergency)",
            "AI Intelligence — anomaly detection, predictive risk, incident correlation",
        ],
        "Show alert firing → AI Chat suggests fix → one-click apply → cluster heals → audit logged"
    ),
    (
        "Change Tasks",
        [
            "Deployment image tag updates",
            "ConfigMap / Secret rotation",
            "HPA scaling policy adjustments",
            "RBAC role binding modifications",
            "Node label & taint management",
            "Operator subscription channel changes",
            "Resource quota adjustments",
        ],
        "40-45%",
        "75-80%",
        [
            "Application Change Watcher — 15s scan cycle, AI impact analysis",
            "Bulk Actions — filter by type/risk, agree/dismiss/acknowledge all",
            "Webhook Notifications — Slack / PagerDuty integration",
            "Workload Timeline — per-resource change history",
            "ConfigMap/Secret Content Tracking — data-level diff detection",
            "MCP Tools — 7 workload management tools (scale, restart, rollback)",
        ],
        "Make a deployment change → watch it appear in 15s → click AI Analysis → see blast radius report"
    ),
    (
        "Change Requests\n(Cluster Upgrades)",
        [
            "End-to-end cluster upgrade lifecycle (10 steps)",
            "22-point pre-assessment (operators, etcd, certs, MCPs, storage)",
            "AI-generated remediation plan with per-fix execution",
            "ServiceNow CR auto-creation with PDF/HTML reports",
            "Approval gate integration (auto-polls every 15s)",
            "Dry run validation before execution",
            "Live CVO monitoring (operator, node, manifest progress)",
            "Post-assessment with pre/post comparison",
            "Auto-close CR with detailed close notes + attached reports",
        ],
        "35-40%",
        "90-95%",
        [
            "Upgrade Orchestrator — 10-step state machine with full audit trail",
            "Pre-Assessment Report — PDF + HTML attached to ServiceNow CR",
            "Post-Assessment Report — verified version, operator/node health, timing",
            "Live Monitoring Dashboard — progress ring, operator donut, CVO manifests",
            "ServiceNow Integration — CR creation, approval gate, auto-closure",
            "Precise Timing — executedAt/completedAt with actual wall-clock duration",
        ],
        "Type 'automate upgrade' → 22 checks run → CR raised → approved → dry run → execute → monitor at 15s intervals → post-assess → CR auto-closed"
    ),
    (
        "SC Tasks\n(Service Catalog)",
        [
            "Namespace provisioning with quotas & RBAC",
            "Service account creation with scoped roles",
            "PVC provisioning with storage class selection",
            "Network policy templates for namespace isolation",
            "Ingress / Route creation for new services",
            "Certificate request & installation",
        ],
        "35-40%",
        "70-75%",
        [
            "MCP Tools — 5 networking tools (ingress, routes, DNS, network policies)",
            "MCP Tools — 4 storage tools (PVC, StorageClass, snapshots)",
            "AI Chat — conversational infrastructure provisioning",
            "GitOps Sync Status — ArgoCD application monitoring",
            "Audit Trail — full compliance logging of all provisioning actions",
        ],
        "Ask AI Chat 'create namespace orders-prod with 4 CPU quota' → resources created → audit logged → GitOps sync verified"
    ),
    (
        "Release Tasks\n(Deployment & Rollout)",
        [
            "Blue-green / canary deployment orchestration",
            "Image vulnerability scanning before deploy",
            "Rollback on failed health checks",
            "GitOps drift detection & auto-sync",
            "DR readiness validation (backups, schedules)",
            "Multi-cluster fleet-wide rollout coordination",
        ],
        "35-40%",
        "70-75%",
        [
            "Image Vulnerabilities Widget — Trivy/ACS scan results in dashboard",
            "GitOps Sync Score — synced/out-of-sync/degraded ArgoCD apps",
            "DR Readiness Score — backup completion, failed jobs, schedule health",
            "App Change Watcher — tracks rollout changes with dismiss confirmation",
            "MCP Tools — 5 GitOps tools (sync, diff, rollback, promote)",
            "Hub/Spoke Architecture — fleet-wide coordination across clusters",
        ],
        "Deploy new image → vuln scan shows results → change watcher detects → AI analyzes impact → GitOps sync confirms → DR backup validates"
    ),
]

# Additional differentiators table
DIFFERENTIATORS = [
    ("Feature", "TCS Agentic AI", "Lens", "Rancher", "K8s Dashboard", "Manual / Scripts"),
    ("AI-Native Operations (LLM-Powered)", "Yes", "No", "No", "No", "No"),
    ("MCP Protocol (44 Tools, Universal Agent)", "Yes", "No", "No", "No", "No"),
    ("Multi-Cluster Hub/Spoke", "Yes", "Limited", "Yes", "No", "Partial"),
    ("End-to-End Upgrade Automation (10 Steps)", "Yes", "No", "No", "No", "Partial"),
    ("ServiceNow CR Integration", "Native", "No", "No", "No", "Manual"),
    ("CIS Compliance Scoring", "Built-in", "Plugin", "External", "No", "External"),
    ("Predictive Risk AI", "Yes", "No", "No", "No", "No"),
    ("Real-Time Change Detection (15s)", "Yes", "No", "No", "No", "No"),
    ("AI Impact Analysis per Change", "Yes", "No", "No", "No", "No"),
    ("Post-Assessment PDF/HTML Reports", "Auto-generated", "No", "No", "No", "Manual"),
    ("Emergency Actions + Audit Trail", "Yes + Audit", "Limited", "Yes", "No", "Manual"),
    ("Natural Language Cluster Operations", "Yes", "No", "No", "No", "No"),
]

# Key numbers
KEY_NUMBERS = [
    ("44", "MCP Tools across 8 categories"),
    ("16", "Real-time dashboard widgets"),
    ("10", "Upgrade orchestration steps (fully automated)"),
    ("22", "Pre-assessment health checks"),
    ("15s", "Change detection scan cycle"),
    ("4", "Integrated views (Dashboard, AI Chat, Audit, AI Intelligence)"),
    ("3", "Governance scorecards (CIS, GitOps, DR)"),
    ("Hub + N", "Multi-cluster spoke architecture"),
]

# ═══════════════════════════════════════════════════════════════════
# EXCEL GENERATION
# ═══════════════════════════════════════════════════════════════════

def generate_excel():
    wb = openpyxl.Workbook()

    # Colors
    NAVY = "0F1B2D"
    BLUE = "1e3a5f"
    LIGHT_BLUE = "DBEAFE"
    WHITE = "FFFFFF"
    GREEN = "15803d"
    GREEN_BG = "D1FAE5"
    ORANGE = "C2410C"
    ORANGE_BG = "FED7AA"
    GRAY = "F9FAFB"
    BORDER_COLOR = "C0C0C0"

    thin_border = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='thin', color=BORDER_COLOR),
    )

    header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
    header_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
    title_font = Font(name="Calibri", bold=True, color=NAVY, size=16)
    subtitle_font = Font(name="Calibri", bold=False, color=BLUE, size=12)
    body_font = Font(name="Calibri", size=10)
    bold_font = Font(name="Calibri", bold=True, size=10)
    pct_font_green = Font(name="Calibri", bold=True, size=12, color=GREEN)
    pct_font_orange = Font(name="Calibri", bold=True, size=12, color=ORANGE)
    wrap_align = Alignment(wrap_text=True, vertical="top")
    center_align = Alignment(wrap_text=True, vertical="center", horizontal="center")

    # ── Sheet 1: Automation Opportunities ──
    ws = wb.active
    ws.title = "Automation Opportunities"
    ws.sheet_properties.tabColor = BLUE

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = TITLE
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:F2")
    ws["A2"] = SUBTITLE
    ws["A2"].font = subtitle_font
    ws.row_dimensions[2].height = 22

    # Work volume reduction header
    ws.merge_cells("D3:E3")
    ws["D3"] = "Work Volume Reduction"
    ws["D3"].font = Font(name="Calibri", bold=True, color=WHITE, size=10)
    ws["D3"].fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    ws["D3"].alignment = center_align

    ws["F3"] = "40-95%"
    ws["F3"].font = Font(name="Calibri", bold=True, color=WHITE, size=10)
    ws["F3"].fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
    ws["F3"].alignment = center_align

    # Column headers
    headers = ["Ticket Type", "Automation Possibilities", "Industry\nBaseline %", "TCS Agentic AI\nAutomation %", "Solution Levers (TCS Agentic AI)", "Key Demo Moment"]
    col_widths = [18, 50, 14, 16, 60, 55]
    row = 5
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Data rows
    for i, (ticket_type, auto_poss, ind_pct, tcs_pct, levers, demo) in enumerate(DATA):
        row = 6 + i
        ws.row_dimensions[row].height = 140

        fill = PatternFill(start_color=(GRAY if i % 2 == 0 else WHITE), end_color=(GRAY if i % 2 == 0 else WHITE), fill_type="solid")

        # Ticket Type
        cell = ws.cell(row=row, column=1, value=ticket_type)
        cell.font = Font(name="Calibri", bold=True, size=11, color=NAVY)
        cell.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = thin_border

        # Automation Possibilities (bullet list)
        cell = ws.cell(row=row, column=2, value="\n".join(f"• {p}" for p in auto_poss))
        cell.font = body_font
        cell.fill = fill
        cell.alignment = wrap_align
        cell.border = thin_border

        # Industry %
        cell = ws.cell(row=row, column=3, value=ind_pct)
        cell.font = pct_font_orange
        cell.fill = PatternFill(start_color=ORANGE_BG, end_color=ORANGE_BG, fill_type="solid")
        cell.alignment = center_align
        cell.border = thin_border

        # TCS %
        cell = ws.cell(row=row, column=4, value=tcs_pct)
        cell.font = pct_font_green
        cell.fill = PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid")
        cell.alignment = center_align
        cell.border = thin_border

        # Solution Levers
        cell = ws.cell(row=row, column=5, value="\n".join(f"✦ {l}" for l in levers))
        cell.font = body_font
        cell.fill = fill
        cell.alignment = wrap_align
        cell.border = thin_border

        # Key Demo Moment
        cell = ws.cell(row=row, column=6, value=demo)
        cell.font = Font(name="Calibri", italic=True, size=10, color=BLUE)
        cell.fill = fill
        cell.alignment = wrap_align
        cell.border = thin_border

    # ── Sheet 2: Competitive Comparison ──
    ws2 = wb.create_sheet("Competitive Comparison")
    ws2.sheet_properties.tabColor = "2563EB"

    ws2.merge_cells("A1:F1")
    ws2["A1"] = "TCS Agentic AI — Competitive Differentiators"
    ws2["A1"].font = title_font
    ws2.row_dimensions[1].height = 35

    for col_idx, h in enumerate(DIFFERENTIATORS[0], 1):
        cell = ws2.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    ws2.column_dimensions["A"].width = 38
    for c in "BCDEF":
        ws2.column_dimensions[c].width = 18

    for row_idx, row_data in enumerate(DIFFERENTIATORS[1:], 4):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = bold_font if col_idx == 1 else body_font
            cell.alignment = center_align if col_idx > 1 else Alignment(wrap_text=True, vertical="center")
            cell.border = thin_border

            if col_idx == 2:  # TCS column
                if val in ("Yes", "Native", "Built-in", "Auto-generated", "Yes + Audit"):
                    cell.fill = PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid")
                    cell.font = Font(name="Calibri", bold=True, size=10, color=GREEN)
            elif col_idx > 2:
                if val == "No":
                    cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                    cell.font = Font(name="Calibri", size=10, color="DC2626")
                elif val in ("Limited", "Partial", "Plugin", "External", "Manual"):
                    cell.fill = PatternFill(start_color=ORANGE_BG, end_color=ORANGE_BG, fill_type="solid")
                    cell.font = Font(name="Calibri", size=10, color=ORANGE)

        ws2.row_dimensions[row_idx].height = 26

    # ── Sheet 3: Key Numbers ──
    ws3 = wb.create_sheet("Key Numbers")
    ws3.sheet_properties.tabColor = GREEN

    ws3.merge_cells("A1:B1")
    ws3["A1"] = "TCS Agentic AI — Key Numbers"
    ws3["A1"].font = title_font
    ws3.row_dimensions[1].height = 35

    for col_idx, h in enumerate(["Metric", "Description"], 1):
        cell = ws3.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    ws3.column_dimensions["A"].width = 15
    ws3.column_dimensions["B"].width = 55

    for row_idx, (num, desc) in enumerate(KEY_NUMBERS, 4):
        cell = ws3.cell(row=row_idx, column=1, value=num)
        cell.font = Font(name="Calibri", bold=True, size=18, color=BLUE)
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")

        cell = ws3.cell(row=row_idx, column=2, value=desc)
        cell.font = Font(name="Calibri", size=12)
        cell.alignment = Alignment(vertical="center")
        cell.border = thin_border
        ws3.row_dimensions[row_idx].height = 32

    # Save
    path = os.path.join(OUT_DIR, "TCS-Agentic-AI-Automation-Opportunities.xlsx")
    wb.save(path)
    print(f"Excel saved: {path}")
    return path


# ═══════════════════════════════════════════════════════════════════
# HTML GENERATION
# ═══════════════════════════════════════════════════════════════════

def generate_html():
    rows_html = ""
    for ticket_type, auto_poss, ind_pct, tcs_pct, levers, demo in DATA:
        auto_list = "".join(f"<li>{p}</li>" for p in auto_poss)
        lever_list = "".join(f"<li>{l}</li>" for l in levers)
        tt = ticket_type.replace("\n", "<br>")
        rows_html += f"""
        <tr>
          <td class="ticket-type">{tt}</td>
          <td class="auto-poss"><ul>{auto_list}</ul></td>
          <td class="pct industry">{ind_pct}</td>
          <td class="pct tcs">{tcs_pct}</td>
          <td class="levers"><ul>{lever_list}</ul></td>
          <td class="demo">{demo}</td>
        </tr>"""

    diff_rows = ""
    for row_data in DIFFERENTIATORS[1:]:
        cells = ""
        for i, val in enumerate(row_data):
            cls = "feature-name" if i == 0 else "tcs-col" if i == 1 else "comp-col"
            badge = ""
            if i == 1 and val in ("Yes", "Native", "Built-in", "Auto-generated", "Yes + Audit"):
                badge = "badge-green"
            elif i > 1 and val == "No":
                badge = "badge-red"
            elif i > 1 and val in ("Limited", "Partial", "Plugin", "External", "Manual"):
                badge = "badge-orange"
            cells += f'<td class="{cls}"><span class="{badge}">{val}</span></td>'
        diff_rows += f"<tr>{cells}</tr>"

    key_nums_html = ""
    for num, desc in KEY_NUMBERS:
        key_nums_html += f"""
        <div class="kn-card">
          <div class="kn-num">{num}</div>
          <div class="kn-desc">{desc}</div>
        </div>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>TCS Agentic AI — Automation Use Cases</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{ font-family: 'Inter', -apple-system, sans-serif; background: #f0f4f8; color: #1e293b; }}

  .page {{ max-width: 1400px; margin: 0 auto; padding: 24px; }}

  /* Header */
  .header {{ background: linear-gradient(135deg, #0f1b2d 0%, #1e3a5f 60%, #2563eb 100%);
    color: #fff; padding: 32px 40px; border-radius: 16px; margin-bottom: 24px;
    display: flex; justify-content: space-between; align-items: center; }}
  .header h1 {{ font-size: 24px; font-weight: 800; letter-spacing: -0.5px; }}
  .header p {{ font-size: 14px; color: rgba(255,255,255,0.7); margin-top: 4px; }}
  .header-badge {{ background: rgba(255,255,255,0.15); border: 1px solid rgba(255,255,255,0.25);
    padding: 8px 20px; border-radius: 10px; text-align: center; }}
  .header-badge .big {{ font-size: 22px; font-weight: 800; }}
  .header-badge .small {{ font-size: 11px; color: rgba(255,255,255,0.7); }}

  /* Key Numbers */
  .kn-grid {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 12px; margin-bottom: 24px; }}
  .kn-card {{ background: #fff; border-radius: 12px; padding: 16px 20px; text-align: center;
    border: 1px solid #e2e8f0; box-shadow: 0 1px 3px rgba(0,0,0,0.06); }}
  .kn-num {{ font-size: 28px; font-weight: 800; color: #1e3a5f; }}
  .kn-desc {{ font-size: 11px; color: #64748b; margin-top: 4px; font-weight: 500; }}

  /* Section */
  .section {{ margin-bottom: 24px; }}
  .section-title {{ font-size: 18px; font-weight: 700; color: #0f1b2d; margin-bottom: 12px;
    padding-bottom: 8px; border-bottom: 2px solid #2563eb; display: inline-block; }}

  /* Main Table */
  .auto-table {{ width: 100%; border-collapse: collapse; background: #fff;
    border-radius: 12px; overflow: hidden; box-shadow: 0 1px 4px rgba(0,0,0,0.08); }}
  .auto-table thead th {{ background: #1e3a5f; color: #fff; padding: 12px 14px;
    font-size: 12px; font-weight: 700; text-align: center; text-transform: uppercase;
    letter-spacing: 0.5px; border-right: 1px solid rgba(255,255,255,0.15); }}
  .auto-table thead th:last-child {{ border-right: none; }}
  .auto-table tbody td {{ padding: 12px 14px; border-bottom: 1px solid #e5e7eb;
    vertical-align: top; font-size: 12px; line-height: 1.5; }}
  .auto-table tbody tr:nth-child(even) {{ background: #f8fafc; }}
  .auto-table tbody tr:hover {{ background: #eff6ff; }}

  .ticket-type {{ font-weight: 700; font-size: 13px; color: #0f1b2d; text-align: center;
    vertical-align: middle !important; background: #dbeafe !important; min-width: 120px; }}
  .auto-poss ul, .levers ul {{ list-style: none; padding: 0; margin: 0; }}
  .auto-poss li {{ padding: 2px 0; }}
  .auto-poss li::before {{ content: "•"; color: #2563eb; font-weight: 700; margin-right: 6px; }}
  .levers li {{ padding: 2px 0; }}
  .levers li::before {{ content: "✦"; color: #15803d; margin-right: 6px; }}
  .pct {{ text-align: center; vertical-align: middle !important; font-size: 16px; font-weight: 800; min-width: 90px; }}
  .pct.industry {{ color: #c2410c; background: #fed7aa !important; }}
  .pct.tcs {{ color: #15803d; background: #d1fae5 !important; }}
  .demo {{ font-style: italic; color: #1e3a5f; font-size: 11px; line-height: 1.6;
    background: #f0f9ff !important; min-width: 200px; }}

  /* Comparison Table */
  .comp-table {{ width: 100%; border-collapse: collapse; background: #fff;
    border-radius: 12px; overflow: hidden; box-shadow: 0 1px 4px rgba(0,0,0,0.08); }}
  .comp-table thead th {{ background: #1e3a5f; color: #fff; padding: 10px 14px;
    font-size: 12px; font-weight: 700; text-align: center; border-right: 1px solid rgba(255,255,255,0.15); }}
  .comp-table tbody td {{ padding: 8px 12px; border-bottom: 1px solid #e5e7eb;
    text-align: center; font-size: 12px; }}
  .comp-table tbody tr:nth-child(even) {{ background: #f8fafc; }}
  .feature-name {{ text-align: left !important; font-weight: 600; min-width: 260px; }}
  .tcs-col {{ font-weight: 700 !important; }}

  .badge-green {{ background: #d1fae5; color: #15803d; padding: 3px 10px;
    border-radius: 6px; font-weight: 700; font-size: 11px; }}
  .badge-red {{ background: #fee2e2; color: #dc2626; padding: 3px 10px;
    border-radius: 6px; font-size: 11px; }}
  .badge-orange {{ background: #fed7aa; color: #c2410c; padding: 3px 10px;
    border-radius: 6px; font-size: 11px; }}

  /* Upgrade Lifecycle */
  .lifecycle {{ display: flex; gap: 0; margin: 16px 0; flex-wrap: wrap; }}
  .lc-step {{ flex: 1; min-width: 100px; padding: 10px 8px; text-align: center;
    background: #1e3a5f; color: #fff; font-size: 10px; font-weight: 600;
    position: relative; clip-path: polygon(0 0, calc(100% - 12px) 0, 100% 50%, calc(100% - 12px) 100%, 0 100%, 12px 50%); }}
  .lc-step:first-child {{ clip-path: polygon(0 0, calc(100% - 12px) 0, 100% 50%, calc(100% - 12px) 100%, 0 100%); }}
  .lc-step:nth-child(odd) {{ background: #2563eb; }}
  .lc-step:nth-child(even) {{ background: #1e3a5f; }}

  .footer {{ text-align: center; margin-top: 24px; padding: 16px;
    color: #94a3b8; font-size: 11px; }}

  @media print {{
    body {{ background: #fff; }}
    .page {{ padding: 0; max-width: 100%; }}
    .header {{ border-radius: 0; }}
    .auto-table, .comp-table {{ box-shadow: none; border: 1px solid #ccc; }}
  }}
</style>
</head>
<body>
<div class="page">

  <div class="header">
    <div>
      <h1>TCS Agentic AI</h1>
      <p>Automation Opportunities for OpenShift Operations — Industry Best Practices</p>
    </div>
    <div style="display:flex;gap:12px;">
      <div class="header-badge">
        <div class="big">44</div>
        <div class="small">MCP Tools</div>
      </div>
      <div class="header-badge">
        <div class="big">10</div>
        <div class="small">Upgrade Steps</div>
      </div>
      <div class="header-badge">
        <div class="big">40-95%</div>
        <div class="small">Work Volume Reduction</div>
      </div>
    </div>
  </div>

  <div class="kn-grid">{key_nums_html}</div>

  <div class="section">
    <div class="section-title">Automation Opportunities by Ticket Type</div>
    <table class="auto-table">
      <thead>
        <tr>
          <th>Ticket Type</th>
          <th>Automation Possibilities</th>
          <th>Industry<br>Baseline</th>
          <th>TCS Agentic AI<br>Automation %</th>
          <th>Solution Levers (TCS Agentic AI)</th>
          <th>Key Demo Moment</th>
        </tr>
      </thead>
      <tbody>{rows_html}</tbody>
    </table>
  </div>

  <div class="section">
    <div class="section-title">End-to-End Cluster Upgrade Lifecycle (10 Automated Steps)</div>
    <div class="lifecycle">
      <div class="lc-step">1. Version<br>Validation</div>
      <div class="lc-step">2. Pre-Assessment<br>(22 Checks)</div>
      <div class="lc-step">3. Component<br>Analysis</div>
      <div class="lc-step">4. Remediation<br>Plan</div>
      <div class="lc-step">5. ServiceNow<br>CR</div>
      <div class="lc-step">6. CR<br>Approval</div>
      <div class="lc-step">7. Dry<br>Run</div>
      <div class="lc-step">8. Execute<br>Upgrade</div>
      <div class="lc-step">9. Live<br>Monitoring</div>
      <div class="lc-step">10. Post-<br>Assessment</div>
    </div>
    <p style="font-size:12px;color:#64748b;margin-top:8px;">
      Each step produces audit trail entries. Pre &amp; post assessment reports (PDF + HTML) are auto-attached to ServiceNow CR.
      CR is auto-closed with verified version, operator/node health, and precise timing (start → complete → duration).
    </p>
  </div>

  <div class="section">
    <div class="section-title">Competitive Differentiators</div>
    <table class="comp-table">
      <thead>
        <tr>
          <th>Feature</th>
          <th>TCS Agentic AI</th>
          <th>Lens</th>
          <th>Rancher</th>
          <th>K8s Dashboard</th>
          <th>Manual / Scripts</th>
        </tr>
      </thead>
      <tbody>{diff_rows}</tbody>
    </table>
  </div>

  <div class="footer">
    Powered by TCS &mdash; Tata Consultancy Services. All rights reserved. &nbsp;|&nbsp; Generated for hackathon demonstration purposes.
  </div>

</div>
</body>
</html>"""

    path = os.path.join(OUT_DIR, "TCS-Agentic-AI-Automation-Opportunities.html")
    with open(path, "w") as f:
        f.write(html)
    print(f"HTML saved: {path}")
    return path


# ═══════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    excel_path = generate_excel()
    html_path = generate_html()
    print(f"\nDone! Files generated:")
    print(f"  Excel: {excel_path}")
    print(f"  HTML:  {html_path}")
