#!/usr/bin/env python3
"""Generate TCS Agentic AI — Upgrade Automation Use Case Excel workbook."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT = os.path.join(os.path.dirname(__file__), "TCS-Agentic-AI-UC02-Cluster-Upgrade.xlsx")

wb = Workbook()

# ── Colours & Styles ──
PRIMARY = "1A1A2E"
ACCENT = "0F7DC2"
GREEN = "16A34A"
RED = "DC2626"
ORANGE = "E67E22"
WHITE = "FFFFFF"
LIGHT_BG = "F8FAFC"
LIGHT_GREEN = "F0FDF4"
LIGHT_RED = "FEF2F2"
LIGHT_BLUE = "EFF6FF"

hdr_font = Font(name="Inter", bold=True, color=WHITE, size=11)
hdr_fill = PatternFill("solid", fgColor=PRIMARY)
sub_hdr_font = Font(name="Inter", bold=True, color=WHITE, size=11)
accent_fill = PatternFill("solid", fgColor=ACCENT)
green_fill = PatternFill("solid", fgColor=GREEN)
red_fill = PatternFill("solid", fgColor=RED)
light_green_fill = PatternFill("solid", fgColor="DCFCE7")
light_red_fill = PatternFill("solid", fgColor="FEE2E2")
light_blue_fill = PatternFill("solid", fgColor="DBEAFE")
light_bg_fill = PatternFill("solid", fgColor=LIGHT_BG)
highlight_fill = PatternFill("solid", fgColor=LIGHT_GREEN)
bold = Font(name="Inter", bold=True, size=11)
bold_green = Font(name="Inter", bold=True, color=GREEN, size=11)
bold_red = Font(name="Inter", bold=True, color=RED, size=11)
bold_white = Font(name="Inter", bold=True, color=WHITE, size=12)
normal = Font(name="Inter", size=11)
small = Font(name="Inter", size=10, color="64748B")
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)


def style_header_row(ws, row, cols, fill=None):
    f = fill or hdr_fill
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font
        cell.fill = f
        cell.alignment = center
        cell.border = thin_border


def style_data_row(ws, row, cols, alt=False, highlight=False):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = normal
        cell.alignment = wrap
        cell.border = thin_border
        if highlight:
            cell.fill = highlight_fill
        elif alt:
            cell.fill = light_bg_fill


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════════════════════════
# SHEET 1: Executive Summary
# ═══════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Executive Summary"
set_col_widths(ws1, [35, 25, 25, 18])

# Title
ws1.merge_cells("A1:D1")
c = ws1["A1"]
c.value = "TCS Agentic AI — End-to-End Cluster Upgrade Automation"
c.font = Font(name="Inter", bold=True, color=WHITE, size=16)
c.fill = PatternFill("solid", fgColor=PRIMARY)
c.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 48

ws1.merge_cells("A2:D2")
c = ws1["A2"]
c.value = "Fully autonomous OpenShift cluster upgrades — from pre-assessment through post-verification — orchestrated by AI with zero manual intervention, complete ITSM integration, and real-time observability."
c.font = Font(name="Inter", size=11, color="64748B", italic=True)
c.fill = light_blue_fill
c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws1.row_dimensions[2].height = 40

# Key Impact Numbers
r = 4
ws1.merge_cells(f"A{r}:D{r}")
ws1.cell(r, 1, "KEY IMPACT METRICS").font = bold
ws1.cell(r, 1).fill = accent_fill
ws1.cell(r, 1).font = bold_white
ws1.cell(r, 1).alignment = center
for c in range(2, 5):
    ws1.cell(r, c).fill = accent_fill

r = 5
metrics = [
    ("Automation Coverage", "90–95%"),
    ("Reduction in Human Effort", "85%"),
    ("Human Touch-Time per Upgrade", "< 10 minutes"),
    ("Audit & Compliance Trail", "100%"),
]
style_header_row(ws1, r, 4, fill=PatternFill("solid", fgColor="0F3460"))
for i, (label, _) in enumerate(metrics, 1):
    ws1.cell(r, i, label)

r = 6
for i, (_, val) in enumerate(metrics, 1):
    cell = ws1.cell(r, i, val)
    cell.font = Font(name="Inter", bold=True, color=GREEN, size=14)
    cell.alignment = center
    cell.fill = light_green_fill
    cell.border = thin_border
ws1.row_dimensions[6].height = 36

# The Challenge
r = 8
ws1.merge_cells(f"A{r}:D{r}")
ws1.cell(r, 1, "THE CHALLENGE: MANUAL CLUSTER UPGRADES").font = bold_white
ws1.cell(r, 1).fill = PatternFill("solid", fgColor=RED)
ws1.cell(r, 1).alignment = center
for c in range(2, 5):
    ws1.cell(r, c).fill = PatternFill("solid", fgColor=RED)

challenges = [
    ("Time-Intensive Process", "Each upgrade requires 14–40 hours of cumulative SRE time spread across 2–5 business days"),
    ("Human Error Risk", "Manual health checks miss edge cases — expired certificates, deprecated APIs, under-replicated etcd"),
    ("Compliance Gaps", "Change Requests often incomplete, post-assessment documentation inconsistent, audit trail gaps"),
    ("Operational Bottleneck", "Senior SREs pulled from high-value work to babysit upgrade progress for hours"),
]
r = 9
ws1.cell(r, 1, "Pain Point").font = bold
ws1.cell(r, 2, "Description").font = bold
ws1.merge_cells(f"B{r}:D{r}")
for c in range(1, 5):
    ws1.cell(r, c).fill = light_red_fill
    ws1.cell(r, c).border = thin_border

for i, (pain, desc) in enumerate(challenges):
    row = 10 + i
    ws1.cell(row, 1, pain).font = Font(name="Inter", bold=True, color=RED, size=11)
    ws1.cell(row, 1).border = thin_border
    ws1.merge_cells(f"B{row}:D{row}")
    ws1.cell(row, 2, desc).font = normal
    ws1.cell(row, 2).alignment = wrap
    ws1.cell(row, 2).border = thin_border

# The Solution
r = 15
ws1.merge_cells(f"A{r}:D{r}")
ws1.cell(r, 1, "THE SOLUTION: TCS AGENTIC AI UPGRADE ORCHESTRATOR").font = bold_white
ws1.cell(r, 1).fill = PatternFill("solid", fgColor=GREEN)
ws1.cell(r, 1).alignment = center
for c in range(2, 5):
    ws1.cell(r, c).fill = PatternFill("solid", fgColor=GREEN)

capabilities = [
    ("AI Pre-Assessment", "Runs 50+ health checks in parallel: operators, nodes, etcd, certs, deprecated APIs, storage, PDBs, networking"),
    ("Intelligent Remediation", "Identifies blockers and proposes targeted fix actions with risk scoring before upgrade begins"),
    ("ITSM Integration", "Auto-generates ServiceNow Change Request with PDF/HTML assessment reports and rollback plan"),
    ("Approval Gate Automation", "Continuously polls ServiceNow for CR approval, auto-advances when approved"),
    ("Safe Execution", "Performs dry-run validation, initiates CVO-driven upgrade with real-time 15-second progress polling"),
    ("Smart Completion", "Upgrade completes only when: ClusterVersion confirmed + all operators healthy + all nodes ready"),
    ("Post-Assessment & Close", "Generates comprehensive report, auto-closes CR with detailed close notes, timing, and metrics"),
]

r = 16
ws1.cell(r, 1, "Capability").font = bold
ws1.cell(r, 2, "Description").font = bold
ws1.merge_cells(f"B{r}:D{r}")
for c in range(1, 5):
    ws1.cell(r, c).fill = light_green_fill
    ws1.cell(r, c).border = thin_border

for i, (cap, desc) in enumerate(capabilities):
    row = 17 + i
    ws1.cell(row, 1, cap).font = Font(name="Inter", bold=True, color=GREEN, size=11)
    ws1.cell(row, 1).border = thin_border
    ws1.cell(row, 1).alignment = wrap
    ws1.merge_cells(f"B{row}:D{row}")
    ws1.cell(row, 2, desc).font = normal
    ws1.cell(row, 2).alignment = wrap
    ws1.cell(row, 2).border = thin_border

# Lifecycle
r = 25
ws1.merge_cells(f"A{r}:D{r}")
ws1.cell(r, 1, "8-PHASE UPGRADE LIFECYCLE").font = bold_white
ws1.cell(r, 1).fill = accent_fill
ws1.cell(r, 1).alignment = center
for c in range(2, 5):
    ws1.cell(r, c).fill = accent_fill

phases = [
    ("01", "Version Validation", "~30 sec", "Validate target version, check upgrade channel, EUS path verification"),
    ("02", "Pre-Assessment", "~2 min", "50+ parallel health checks across operators, nodes, etcd, certs, APIs, storage"),
    ("03", "Component Analysis", "~1 min", "Identify blockers, risk-score components, generate remediation proposals"),
    ("04", "CR Submission", "~30 sec", "Auto-create ServiceNow CR with PDF/HTML attachments, risk level, rollback plan"),
    ("05", "Approval Gate", "Auto-poll", "Continuously poll ServiceNow for approval, auto-advance when approved"),
    ("06", "Dry Run", "~1 min", "Validate upgrade path, preflight checks, verify no new blockers since assessment"),
    ("07", "Execute & Monitor", "Cluster-paced", "Initiate CVO upgrade, 15-second polling, operator/node health tracking"),
    ("08", "Post-Assessment", "~2 min", "Generate final report with timing, operator/node counts, auto-close CR"),
]

r = 26
headers = ["Phase", "Step", "Duration", "Description"]
style_header_row(ws1, r, 4)
for i, h in enumerate(headers, 1):
    ws1.cell(r, i, h)

for i, (num, step, dur, desc) in enumerate(phases):
    row = 27 + i
    ws1.cell(row, 1, num).font = bold
    ws1.cell(row, 1).alignment = center
    ws1.cell(row, 1).border = thin_border
    ws1.cell(row, 2, step).font = bold
    ws1.cell(row, 2).border = thin_border
    ws1.cell(row, 3, dur).font = bold_green
    ws1.cell(row, 3).alignment = center
    ws1.cell(row, 3).border = thin_border
    ws1.cell(row, 4, desc).font = normal
    ws1.cell(row, 4).alignment = wrap
    ws1.cell(row, 4).border = thin_border
    if i % 2 == 1:
        for c in range(1, 5):
            ws1.cell(row, c).fill = light_bg_fill


# ═══════════════════════════════════════════════════════════════════
# SHEET 2: Manual vs Automated
# ═══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Manual vs Automated")
set_col_widths(ws2, [28, 22, 22, 14, 30])

ws2.merge_cells("A1:E1")
c = ws2["A1"]
c.value = "Side-by-Side Comparison: Manual vs. TCS Agentic AI Automated"
c.font = Font(name="Inter", bold=True, color=WHITE, size=14)
c.fill = hdr_fill
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 40

r = 3
headers = ["Upgrade Phase", "Manual (SRE Time)", "Automated (Human Time)", "Time Saved", "Automation Method"]
style_header_row(ws2, r, 5)
for i, h in enumerate(headers, 1):
    ws2.cell(r, i, h)

comparison = [
    ("Version Validation", "15–30 min", "~30 sec", "95%", "AI — Channel graph analysis, EUS detection"),
    ("Pre-Assessment (50+ checks)", "2–4 hours", "2–3 min", "97%", "AI — Parallel health check execution"),
    ("Remediation Planning", "1–3 hours", "1 min (review)", "95%", "AI — Risk-scored fix proposals"),
    ("Change Request Creation", "1–2 hours", "~30 sec", "96%", "AUTO — ServiceNow API with PDF/HTML"),
    ("Approval Wait & Follow-up", "4–24 hours (calendar)", "0 min (auto-poll)", "100%", "AUTO — Continuous ServiceNow polling"),
    ("Dry Run Validation", "30–60 min", "1–2 min", "95%", "AUTO — Automated preflight checks"),
    ("Upgrade Execution", "2–4 hours (active monitoring)", "0 min (autonomous)", "100%", "AUTO — CVO-driven with 15s polling"),
    ("Post-Upgrade Verification", "2–4 hours", "0 min (smart completion)", "100%", "AI — Triple-condition verification"),
    ("Post-Assessment Report", "1–2 hours", "1–2 min", "96%", "AI — PDF/HTML with timing & metrics"),
    ("CR Closure & Documentation", "30–60 min", "~30 sec", "97%", "AUTO — Auto-close with detailed notes"),
]

for i, (phase, manual, auto, saved, method) in enumerate(comparison):
    row = 4 + i
    ws2.cell(row, 1, phase).font = bold
    ws2.cell(row, 1).border = thin_border
    ws2.cell(row, 1).alignment = wrap
    ws2.cell(row, 2, manual).font = Font(name="Inter", color=RED, size=11)
    ws2.cell(row, 2).alignment = center
    ws2.cell(row, 2).border = thin_border
    ws2.cell(row, 3, auto).font = Font(name="Inter", color=GREEN, size=11)
    ws2.cell(row, 3).alignment = center
    ws2.cell(row, 3).border = thin_border
    ws2.cell(row, 4, saved).font = bold_green
    ws2.cell(row, 4).alignment = center
    ws2.cell(row, 4).border = thin_border
    ws2.cell(row, 5, method).font = normal
    ws2.cell(row, 5).alignment = wrap
    ws2.cell(row, 5).border = thin_border
    if i % 2 == 1:
        for c in range(1, 6):
            ws2.cell(row, c).fill = light_bg_fill

# Total row
row = 14
ws2.cell(row, 1, "TOTAL (per cluster)").font = Font(name="Inter", bold=True, color=WHITE, size=12)
ws2.cell(row, 1).fill = PatternFill("solid", fgColor=PRIMARY)
ws2.cell(row, 1).border = thin_border
ws2.cell(row, 2, "14–40 hours").font = Font(name="Inter", bold=True, color=WHITE, size=12)
ws2.cell(row, 2).fill = PatternFill("solid", fgColor=RED)
ws2.cell(row, 2).alignment = center
ws2.cell(row, 2).border = thin_border
ws2.cell(row, 3, "< 10 minutes").font = Font(name="Inter", bold=True, color=WHITE, size=12)
ws2.cell(row, 3).fill = PatternFill("solid", fgColor=GREEN)
ws2.cell(row, 3).alignment = center
ws2.cell(row, 3).border = thin_border
ws2.cell(row, 4, "~95%").font = Font(name="Inter", bold=True, color=WHITE, size=12)
ws2.cell(row, 4).fill = PatternFill("solid", fgColor=GREEN)
ws2.cell(row, 4).alignment = center
ws2.cell(row, 4).border = thin_border
ws2.cell(row, 5, "End-to-End AI Orchestration").font = Font(name="Inter", bold=True, color=WHITE, size=11)
ws2.cell(row, 5).fill = PatternFill("solid", fgColor=PRIMARY)
ws2.cell(row, 5).alignment = center
ws2.cell(row, 5).border = thin_border
ws2.row_dimensions[row].height = 32

# Multi-Cluster Scale
r = 17
ws2.merge_cells(f"A{r}:E{r}")
ws2.cell(r, 1, "MULTI-CLUSTER IMPACT AT SCALE").font = bold_white
ws2.cell(r, 1).fill = accent_fill
ws2.cell(r, 1).alignment = center
for c in range(2, 6):
    ws2.cell(r, c).fill = accent_fill

r = 18
for i, h in enumerate(["Environment Scale", "Clusters", "Manual Effort (per cycle)", "With TCS Agentic AI", "SRE Hours Reclaimed"], 1):
    ws2.cell(r, i, h)
style_header_row(ws2, r, 5)

scale_data = [
    ("Small", "5", "70–200 hours", "~50 min human + cluster time", "~68–198 hours"),
    ("Medium", "15", "210–600 hours", "~2.5 hrs human + cluster time", "~207–597 hours"),
    ("Enterprise", "50", "700–2,000 hours", "~8 hrs human + cluster time", "~692–1,992 hours"),
]

for i, (scale, clusters, manual, auto, saved) in enumerate(scale_data):
    row = 19 + i
    ws2.cell(row, 1, scale).font = bold
    ws2.cell(row, 1).border = thin_border
    ws2.cell(row, 2, clusters).font = bold
    ws2.cell(row, 2).alignment = center
    ws2.cell(row, 2).border = thin_border
    ws2.cell(row, 3, manual).font = bold_red
    ws2.cell(row, 3).alignment = center
    ws2.cell(row, 3).border = thin_border
    ws2.cell(row, 4, auto).font = bold_green
    ws2.cell(row, 4).alignment = center
    ws2.cell(row, 4).border = thin_border
    ws2.cell(row, 5, saved).font = bold_green
    ws2.cell(row, 5).alignment = center
    ws2.cell(row, 5).border = thin_border
    if i % 2 == 1:
        for c in range(1, 6):
            ws2.cell(row, c).fill = light_bg_fill


# ═══════════════════════════════════════════════════════════════════
# SHEET 3: Health Checks
# ═══════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Pre-Assessment Checks")
set_col_widths(ws3, [22, 40, 16, 30])

ws3.merge_cells("A1:D1")
c = ws3["A1"]
c.value = "AI Pre-Assessment: 50+ Automated Health Checks"
c.font = Font(name="Inter", bold=True, color=WHITE, size=14)
c.fill = hdr_fill
c.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 40

r = 3
headers = ["Category", "Checks Performed", "Mode", "Blocker Detection"]
style_header_row(ws3, r, 4)
for i, h in enumerate(headers, 1):
    ws3.cell(r, i, h)

checks = [
    ("Cluster Operators", "All ClusterOperators available, not degraded, not progressing; version consistency across all operators", "AI Analysis", "Degraded or unavailable operators block upgrade"),
    ("Node Health", "All nodes Ready, no SchedulingDisabled, resource pressure (memory/disk/PID), kernel version compatibility", "Automated", "NotReady or cordoned nodes flagged"),
    ("etcd Cluster", "Quorum health, member count, leader stability, DB size, defragmentation status", "AI Analysis", "Quorum loss or oversized DB blocks upgrade"),
    ("Certificate Expiry", "All cluster certificates checked for expiry within upgrade window; kube-apiserver, etcd, ingress certs", "Automated", "Certificates expiring <30 days flagged"),
    ("Deprecated APIs", "Scan for removed/deprecated Kubernetes APIs in target version; workload manifest analysis", "AI Analysis", "Removed APIs in use block upgrade"),
    ("Storage & PVs", "PersistentVolume capacity, StorageClass health, CSI driver status, volume attachment limits", "Automated", "Full PVs or failed CSI drivers flagged"),
    ("Pod Disruption Budgets", "PDB configuration validation, max unavailable settings, critical workload protection", "AI Analysis", "Misconfigured PDBs flagged with fix proposals"),
    ("Networking", "SDN/OVN health, DNS resolution, ingress controller status, service mesh compatibility", "Automated", "Network operator issues block upgrade"),
    ("Upgrade Path", "Channel graph validation, EUS-to-EUS path verification, intermediate version requirements", "AI Analysis", "Invalid upgrade path blocks immediately"),
    ("Machine Config", "MachineConfigPool sync status, pending machine configs, paused pools", "Automated", "Paused or degraded MCPs flagged"),
]

for i, (cat, checks_desc, mode, blocker) in enumerate(checks):
    row = 4 + i
    ws3.cell(row, 1, cat).font = bold
    ws3.cell(row, 1).border = thin_border
    ws3.cell(row, 1).alignment = wrap
    ws3.cell(row, 2, checks_desc).font = normal
    ws3.cell(row, 2).alignment = wrap
    ws3.cell(row, 2).border = thin_border
    cell_mode = ws3.cell(row, 3, mode)
    cell_mode.alignment = center
    cell_mode.border = thin_border
    if "AI" in mode:
        cell_mode.font = Font(name="Inter", bold=True, color="1E40AF", size=10)
        cell_mode.fill = light_blue_fill
    else:
        cell_mode.font = Font(name="Inter", bold=True, color=GREEN, size=10)
        cell_mode.fill = light_green_fill
    ws3.cell(row, 4, blocker).font = normal
    ws3.cell(row, 4).alignment = wrap
    ws3.cell(row, 4).border = thin_border
    if i % 2 == 1:
        ws3.cell(row, 1).fill = light_bg_fill
        ws3.cell(row, 2).fill = light_bg_fill
        ws3.cell(row, 4).fill = light_bg_fill


# ═══════════════════════════════════════════════════════════════════
# SHEET 4: Real-World Scenarios
# ═══════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Upgrade Scenarios")
set_col_widths(ws4, [28, 24, 20, 24, 34])

ws4.merge_cells("A1:E1")
c = ws4["A1"]
c.value = "Real-World Upgrade Scenarios"
c.font = Font(name="Inter", bold=True, color=WHITE, size=14)
c.fill = hdr_fill
c.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 40

r = 3
headers = ["Scenario", "Cluster Profile", "Manual Duration", "With TCS Agentic AI", "Outcome"]
style_header_row(ws4, r, 5)
for i, h in enumerate(headers, 1):
    ws4.cell(r, i, h)

scenarios = [
    ("Minor Version Upgrade\n4.14.x → 4.14.y", "3 masters, 6 workers\nProduction", "~16 hours\n(2 days)", "~8 min human\n+ 45 min cluster", "Zero blocker misses, CR auto-closed with full report"),
    ("EUS-to-EUS Upgrade\n4.12 → 4.14 (via 4.13)", "3 masters, 12 workers\nProduction", "~32 hours\n(3–5 days)", "~10 min human\n+ 2.5 hrs cluster", "Multi-hop path auto-validated, deprecated APIs caught pre-upgrade"),
    ("Fleet Upgrade\n15 clusters, rolling", "Mixed sizes\nDev + Staging + Prod", "~300 hours\n(2–3 weeks)", "~2.5 hrs human\n+ cluster time", "Rolling upgrade orchestration, per-cluster reports, fleet-wide compliance"),
    ("Emergency Security Patch\nCVE remediation", "5 clusters\nCritical production", "~40 hours\n(expedited)", "~50 min human\n+ cluster time", "Fast-track assessment, expedited CR, minimal approval delay"),
]

for i, (scenario, profile, manual, auto, outcome) in enumerate(scenarios):
    row = 4 + i
    ws4.cell(row, 1, scenario).font = bold
    ws4.cell(row, 1).alignment = wrap
    ws4.cell(row, 1).border = thin_border
    ws4.cell(row, 2, profile).font = normal
    ws4.cell(row, 2).alignment = wrap
    ws4.cell(row, 2).border = thin_border
    ws4.cell(row, 3, manual).font = bold_red
    ws4.cell(row, 3).alignment = center
    ws4.cell(row, 3).border = thin_border
    ws4.cell(row, 4, auto).font = bold_green
    ws4.cell(row, 4).alignment = center
    ws4.cell(row, 4).border = thin_border
    ws4.cell(row, 5, outcome).font = normal
    ws4.cell(row, 5).alignment = wrap
    ws4.cell(row, 5).border = thin_border
    ws4.row_dimensions[row].height = 44
    if i % 2 == 1:
        for c in range(1, 6):
            ws4.cell(row, c).fill = light_bg_fill


# ═══════════════════════════════════════════════════════════════════
# SHEET 5: Safety & Risk
# ═══════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Safety & Risk Mitigation")
set_col_widths(ws5, [24, 32, 40])

ws5.merge_cells("A1:C1")
c = ws5["A1"]
c.value = "Built-in Safety & Risk Mitigation"
c.font = Font(name="Inter", bold=True, color=WHITE, size=14)
c.fill = hdr_fill
c.alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[1].height = 40

r = 3
headers = ["Risk Scenario", "Description", "Built-in Mitigation"]
style_header_row(ws5, r, 3)
for i, h in enumerate(headers, 1):
    ws5.cell(r, i, h)

risks = [
    ("Stalled Upgrade", "CVO progress stuck at a percentage for extended time", "15-second polling detects stall, alerts SRE, provides operator-level diagnostics"),
    ("Operator Degradation", "ClusterOperator enters degraded state during upgrade", "Real-time operator health tracking, AI analysis of degradation cause, rollback guidance"),
    ("Node Not Ready", "Worker or master node fails to rejoin after reboot", "Node-level health monitoring, machine config pool status tracking, drain/cordon detection"),
    ("False Completion", "Upgrade appears done but operators still reconciling", "Triple-condition gate: ClusterVersion history confirmed + all operators healthy + all nodes ready"),
    ("Approval Delay", "ServiceNow CR approval takes longer than expected", "Continuous auto-polling, state machine pauses safely at approval gate without timeout"),
    ("Pre-Assessment Miss", "Edge-case health issue not caught before upgrade", "50+ parallel checks covering operators, nodes, etcd, certs, APIs, storage, PDBs, networking"),
]

for i, (risk, desc, mitigation) in enumerate(risks):
    row = 4 + i
    ws5.cell(row, 1, risk).font = Font(name="Inter", bold=True, color=RED, size=11)
    ws5.cell(row, 1).border = thin_border
    ws5.cell(row, 1).alignment = wrap
    ws5.cell(row, 2, desc).font = normal
    ws5.cell(row, 2).alignment = wrap
    ws5.cell(row, 2).border = thin_border
    ws5.cell(row, 3, mitigation).font = Font(name="Inter", bold=True, color=GREEN, size=11)
    ws5.cell(row, 3).alignment = wrap
    ws5.cell(row, 3).border = thin_border
    if i % 2 == 1:
        for c in range(1, 4):
            ws5.cell(row, c).fill = light_bg_fill


# ═══════════════════════════════════════════════════════════════════
# SHEET 6: Key Benefits
# ═══════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Key Benefits")
set_col_widths(ws6, [28, 50])

ws6.merge_cells("A1:B1")
c = ws6["A1"]
c.value = "Key Benefits of TCS Agentic AI Upgrade Automation"
c.font = Font(name="Inter", bold=True, color=WHITE, size=14)
c.fill = hdr_fill
c.alignment = Alignment(horizontal="center", vertical="center")
ws6.row_dimensions[1].height = 40

r = 3
style_header_row(ws6, r, 2)
ws6.cell(r, 1, "Benefit")
ws6.cell(r, 2, "Description")

benefits = [
    ("95% Time Reduction", "From 14–40 hours of SRE effort to under 10 minutes of human touch-time per cluster upgrade"),
    ("Zero-Touch Execution", "Autonomous monitoring with 15-second polling, smart triple-condition completion detection — no human babysitting"),
    ("Complete Audit Trail", "Every step timestamped and documented — PDF reports, ServiceNow CRs, work notes, and detailed close notes"),
    ("Intelligent Remediation", "AI identifies upgrade blockers before they cause failures and proposes risk-scored fix actions with one-click apply"),
    ("Multi-Cluster Scale", "Hub-spoke architecture manages upgrades across 50+ clusters from a single control plane with per-cluster tracking"),
    ("ITSM Compliance", "Full ServiceNow integration — auto-create CR, approval gate polling, work notes updates, and auto-close with evidence"),
    ("Real-Time Observability", "Live dashboard with CVO percentage, operator health matrix, node readiness, and timing metrics during upgrade"),
    ("Safe Rollback Path", "Rollback plan auto-generated and attached to CR; state machine can pause at any gate for manual intervention"),
]

for i, (benefit, desc) in enumerate(benefits):
    row = 4 + i
    ws6.cell(row, 1, benefit).font = bold_green
    ws6.cell(row, 1).border = thin_border
    ws6.cell(row, 1).alignment = wrap
    ws6.cell(row, 2, desc).font = normal
    ws6.cell(row, 2).alignment = wrap
    ws6.cell(row, 2).border = thin_border
    if i % 2 == 1:
        for c in range(1, 3):
            ws6.cell(row, c).fill = light_bg_fill


# ── Save ──
wb.save(OUT)
print(f"Excel saved: {OUT}")
