#!/usr/bin/env python3
"""Generate TCS Agentic AI — Upgrade Use Case & Workflow Excel workbook."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT = os.path.join(os.path.dirname(__file__), "TCS-Agentic-AI-UC02-Cluster-Upgrade-Workflow.xlsx")
wb = Workbook()

# Styles
PRIMARY = "1A1A2E"
ACCENT = "0F7DC2"
GREEN = "16A34A"
RED = "DC2626"
WHITE = "FFFFFF"

hdr_font = Font(name="Inter", bold=True, color=WHITE, size=11)
hdr_fill = PatternFill("solid", fgColor=PRIMARY)
accent_fill = PatternFill("solid", fgColor=ACCENT)
green_fill = PatternFill("solid", fgColor=GREEN)
red_fill = PatternFill("solid", fgColor=RED)
light_green = PatternFill("solid", fgColor="DCFCE7")
light_red = PatternFill("solid", fgColor="FEE2E2")
light_blue = PatternFill("solid", fgColor="DBEAFE")
light_bg = PatternFill("solid", fgColor="F8FAFC")
bold = Font(name="Inter", bold=True, size=11)
bold_w = Font(name="Inter", bold=True, color=WHITE, size=11)
bold_w_lg = Font(name="Inter", bold=True, color=WHITE, size=14)
bold_green = Font(name="Inter", bold=True, color=GREEN, size=11)
bold_red = Font(name="Inter", bold=True, color=RED, size=11)
bold_accent = Font(name="Inter", bold=True, color=ACCENT, size=11)
normal = Font(name="Inter", size=11)
small_gray = Font(name="Inter", size=10, color="64748B")
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Border(
    left=Side("thin", color="E2E8F0"), right=Side("thin", color="E2E8F0"),
    top=Side("thin", color="E2E8F0"), bottom=Side("thin", color="E2E8F0"),
)

def hdr_row(ws, r, cols, fill=None):
    for c in range(1, cols + 1):
        cell = ws.cell(r, c)
        cell.font = hdr_font
        cell.fill = fill or hdr_fill
        cell.alignment = center
        cell.border = thin

def col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════════════════════════
# SHEET 1: Use Case & Time Comparison
# ═══════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Use Case & Time Comparison"
col_widths(ws1, [6, 34, 22, 24, 12, 14])

# Title
ws1.merge_cells("A1:F1")
c = ws1["A1"]
c.value = "TCS Agentic AI — End-to-End Cluster Upgrade Automation"
c.font = Font(name="Inter", bold=True, color=WHITE, size=16)
c.fill = hdr_fill
c.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 44

# Use Case Box
ws1.merge_cells("A3:F3")
ws1.cell(3, 1, "USE CASE").font = bold_w
ws1.cell(3, 1).fill = accent_fill
ws1.cell(3, 1).alignment = center
for c in range(2, 7):
    ws1.cell(3, c).fill = accent_fill

ws1.merge_cells("A4:B4")
ws1.cell(4, 1, "Use Case Title:").font = bold
ws1.cell(4, 1).border = thin
ws1.merge_cells("C4:F4")
ws1.cell(4, 3, "Autonomous OpenShift Cluster Version Upgrade").font = Font(name="Inter", bold=True, size=13, color=PRIMARY)
ws1.cell(4, 3).border = thin
ws1.row_dimensions[4].height = 28

ws1.merge_cells("A5:B5")
ws1.cell(5, 1, "Description:").font = bold
ws1.cell(5, 1).border = thin
ws1.cell(5, 1).alignment = wrap
ws1.merge_cells("C5:F5")
ws1.cell(5, 3).value = (
    'An SRE or Platform Engineer initiates a cluster upgrade through the AI Chat interface by simply '
    'stating the target version (e.g., "Upgrade this cluster to 4.14.28"). TCS Agentic AI autonomously '
    'orchestrates the entire upgrade lifecycle — pre-flight health assessment, ITSM change management, '
    'safe execution with real-time monitoring, and post-upgrade verification with automated documentation — '
    'reducing 14–40 hours of manual SRE effort to under 10 minutes of human involvement per cluster.'
)
ws1.cell(5, 3).font = normal
ws1.cell(5, 3).alignment = wrap
ws1.cell(5, 3).border = thin
ws1.row_dimensions[5].height = 72

# Key Numbers
ws1.merge_cells("A7:F7")
ws1.cell(7, 1, "KEY METRICS").font = bold_w
ws1.cell(7, 1).fill = PatternFill("solid", fgColor=GREEN)
ws1.cell(7, 1).alignment = center
for c in range(2, 7):
    ws1.cell(7, c).fill = PatternFill("solid", fgColor=GREEN)

metrics = [
    ("Manual Effort (per cluster)", "14–40 hours", bold_red),
    ("Human Time with Agentic AI", "< 10 minutes", bold_green),
    ("Time Saved", "~95%", bold_green),
    ("Workflow Steps (Fully Automated)", "13 Steps", bold_accent),
    ("Automation Coverage", "90–95%", bold_green),
    ("Audit & Compliance Trail", "100%", bold_green),
]
for i, (label, val, vfont) in enumerate(metrics):
    r = 8 + i
    ws1.merge_cells(f"A{r}:C{r}")
    ws1.cell(r, 1, label).font = bold
    ws1.cell(r, 1).border = thin
    ws1.cell(r, 1).alignment = wrap
    ws1.merge_cells(f"D{r}:F{r}")
    ws1.cell(r, 4, val).font = vfont
    ws1.cell(r, 4).border = thin
    ws1.cell(r, 4).alignment = center
    if i % 2 == 0:
        for c2 in [1, 4]:
            ws1.cell(r, c2).fill = light_bg

# Time Comparison Table
r = 16
ws1.merge_cells(f"A{r}:F{r}")
ws1.cell(r, 1, "MANUAL PROCESS vs. TCS AGENTIC AI — TIME COMPARISON").font = bold_w
ws1.cell(r, 1).fill = hdr_fill
ws1.cell(r, 1).alignment = center
for c in range(2, 7):
    ws1.cell(r, c).fill = hdr_fill

r = 17
headers = ["#", "Activity", "Manual (SRE Time)", "Automated (Agentic AI)", "Saved", "Mode"]
hdr_row(ws1, r, 6)
for i, h in enumerate(headers, 1):
    ws1.cell(r, i, h)
ws1.cell(r, 3).fill = PatternFill("solid", fgColor=RED)
ws1.cell(r, 4).fill = PatternFill("solid", fgColor=GREEN)

rows = [
    ("1", "Identify target version & validate upgrade path", "15–30 min", "~30 sec", "95%", "AI"),
    ("2", "Pre-upgrade health checks\n(Operators, nodes, etcd, certs, APIs, storage, PDBs)", "2–4 hours", "2–3 min", "97%", "AI"),
    ("3", "Blocker identification & remediation planning", "1–3 hours", "~1 min", "95%", "AI"),
    ("4", "Create ServiceNow Change Request\n(Justification, risk, rollback plan, attachments)", "1–2 hours", "~30 sec", "96%", "AUTO"),
    ("5", "Wait for CAB / manager approval", "4–24 hours", "0 min (auto-poll)", "100%", "AUTO"),
    ("6", "Dry-run validation & preflight checks", "30–60 min", "1–2 min", "95%", "AUTO"),
    ("7", "Execute upgrade & active monitoring\n(Watch CVO progress, operators, nodes for hours)", "2–4 hours", "0 min (autonomous)", "100%", "AUTO"),
    ("8", "Post-upgrade verification & smoke tests", "2–4 hours", "0 min (smart detect)", "100%", "AI"),
    ("9", "Generate post-assessment report (PDF/HTML)", "1–2 hours", "1–2 min", "96%", "AI"),
    ("10", "Close Change Request with documentation", "30–60 min", "~30 sec", "97%", "AUTO"),
]

for i, (num, act, manual, auto, saved, mode) in enumerate(rows):
    r = 18 + i
    ws1.cell(r, 1, num).font = bold
    ws1.cell(r, 1).alignment = center
    ws1.cell(r, 1).border = thin
    ws1.cell(r, 2, act).font = normal
    ws1.cell(r, 2).alignment = wrap
    ws1.cell(r, 2).border = thin
    ws1.cell(r, 3, manual).font = bold_red
    ws1.cell(r, 3).alignment = center
    ws1.cell(r, 3).border = thin
    ws1.cell(r, 4, auto).font = bold_green
    ws1.cell(r, 4).alignment = center
    ws1.cell(r, 4).border = thin
    ws1.cell(r, 5, saved).font = bold_green
    ws1.cell(r, 5).alignment = center
    ws1.cell(r, 5).border = thin

    mode_cell = ws1.cell(r, 6, mode)
    mode_cell.alignment = center
    mode_cell.border = thin
    if mode == "AI":
        mode_cell.font = Font(name="Inter", bold=True, color="1E40AF", size=11)
        mode_cell.fill = light_blue
    else:
        mode_cell.font = Font(name="Inter", bold=True, color=GREEN, size=11)
        mode_cell.fill = light_green

    if "\n" in act:
        ws1.row_dimensions[r].height = 36
    if i % 2 == 1:
        ws1.cell(r, 1).fill = light_bg
        ws1.cell(r, 2).fill = light_bg
        ws1.cell(r, 3).fill = light_bg
        ws1.cell(r, 4).fill = light_bg
        ws1.cell(r, 5).fill = light_bg

# Total
r = 28
ws1.cell(r, 1, "").fill = hdr_fill
ws1.cell(r, 1).border = thin
ws1.cell(r, 2, "TOTAL PER CLUSTER").font = Font(name="Inter", bold=True, color=WHITE, size=12)
ws1.cell(r, 2).fill = hdr_fill
ws1.cell(r, 2).border = thin
ws1.cell(r, 3, "14–40 hours").font = Font(name="Inter", bold=True, color=WHITE, size=13)
ws1.cell(r, 3).fill = red_fill
ws1.cell(r, 3).alignment = center
ws1.cell(r, 3).border = thin
ws1.cell(r, 4, "< 10 minutes").font = Font(name="Inter", bold=True, color=WHITE, size=13)
ws1.cell(r, 4).fill = green_fill
ws1.cell(r, 4).alignment = center
ws1.cell(r, 4).border = thin
ws1.cell(r, 5, "~95%").font = Font(name="Inter", bold=True, color=WHITE, size=13)
ws1.cell(r, 5).fill = green_fill
ws1.cell(r, 5).alignment = center
ws1.cell(r, 5).border = thin
ws1.cell(r, 6, "").fill = hdr_fill
ws1.cell(r, 6).border = thin
ws1.row_dimensions[r].height = 32


# ═══════════════════════════════════════════════════════════════════
# SHEET 2: Step-by-Step Workflow
# ═══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Upgrade Workflow (13 Steps)")
col_widths(ws2, [8, 30, 12, 50, 14])

ws2.merge_cells("A1:E1")
c = ws2["A1"]
c.value = "Automated Upgrade Workflow — Step by Step"
c.font = Font(name="Inter", bold=True, color=WHITE, size=14)
c.fill = hdr_fill
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 40

ws2.merge_cells("A2:E2")
ws2.cell(2, 1).value = 'When a user types "Upgrade this cluster to 4.14.28" in the AI Chat, the following 13-step workflow executes automatically:'
ws2.cell(2, 1).font = Font(name="Inter", size=11, color="64748B", italic=True)
ws2.cell(2, 1).alignment = wrap
ws2.row_dimensions[2].height = 28

r = 4
headers = ["Step", "Title", "Mode", "What Happens", "Duration"]
hdr_row(ws2, r, 5)
for i, h in enumerate(headers, 1):
    ws2.cell(r, i, h)

workflow = [
    ("1", "User Request & Intent Recognition", "USER",
     "User types a natural language upgrade request in AI Chat.\n"
     "- AI parses message and identifies upgrade intent and target version\n"
     "- Extracts version number (e.g., 4.14.28) from free-form text\n"
     "- Confirms cluster context from active workspace selection",
     "Instant"),

    ("2", "Version Validation & Path Verification", "AI",
     "AI validates the target version and upgrade path.\n"
     "- Queries ClusterVersion to get current version\n"
     "- Checks upgrade channel graph for valid paths\n"
     "- Detects EUS-to-EUS upgrades and validates intermediate hops\n"
     "- Confirms target version is available in configured channel",
     "~30 sec"),

    ("3", "Pre-Assessment — 50+ Health Checks", "AI",
     "AI runs comprehensive cluster health assessment in parallel.\n"
     "- Cluster Operators: all available, not degraded, not progressing\n"
     "- Node Health: all Ready, no resource pressure, no cordoned nodes\n"
     "- etcd: quorum, member count, leader stability, DB size\n"
     "- Certificates: check expiry within upgrade window\n"
     "- Deprecated APIs: scan for removed APIs in target version\n"
     "- Storage: PV capacity, CSI drivers, StorageClass health\n"
     "- Pod Disruption Budgets: validate PDB configuration\n"
     "- Networking: SDN/OVN, DNS, ingress controllers\n"
     "- Machine Config Pools: sync status, paused pools",
     "2–3 min"),

    ("4", "Component Analysis & Blocker Identification", "AI",
     "AI analyzes health check results and identifies blockers.\n"
     "- Classifies each finding as Blocker, Warning, or Info\n"
     "- Assigns risk scores to each component\n"
     "- Generates overall upgrade readiness score\n"
     "- Produces PDF and HTML assessment reports",
     "~1 min"),

    ("5", "Remediation Proposals", "AI",
     "For any blockers found, AI proposes targeted fix actions.\n"
     "- Each blocker gets a specific remediation action\n"
     "- Fix proposals include risk level and impact assessment\n"
     "- User can review and approve fixes via the UI\n"
     "- If no blockers found, this step auto-advances",
     "~30 sec"),

    ("6", "ServiceNow Change Request Creation", "AUTO",
     "System auto-generates a fully populated ServiceNow CR.\n"
     "- CR populated with: description, justification, risk classification, rollback plan\n"
     "- PDF assessment report attached as evidence\n"
     "- HTML assessment report attached for easy viewing\n"
     "- CR category, priority, and assignment group auto-set",
     "~30 sec"),

    ("7", "Approval Gate — Auto-Polling", "GATE",
     "System continuously polls ServiceNow for CR approval.\n"
     "- Polls ServiceNow API every 60 seconds for state changes\n"
     "- Auto-advances when CR reaches 'Approved' state\n"
     "- State machine pauses safely — no timeout, no data loss\n"
     "- Dashboard shows live CR status to the user",
     "Auto-poll"),

    ("8", "Dry Run Validation", "AUTO",
     "Final preflight validation before committing to upgrade.\n"
     "- Re-validates upgrade path is still available\n"
     "- Checks no new blockers appeared since pre-assessment\n"
     "- Verifies cluster state hasn't degraded during approval wait\n"
     "- Confirms all prerequisites still met",
     "1–2 min"),

    ("9", "Execute Upgrade", "AUTO",
     "System initiates the CVO-driven cluster upgrade.\n"
     "- Patches ClusterVersion resource with target version\n"
     "- Records precise executedAt timestamp for audit trail\n"
     "- Updates ServiceNow CR with work notes: 'Upgrade initiated'\n"
     "- Transitions state machine to monitoring phase",
     "~30 sec"),

    ("10", "Real-Time Progress Monitoring", "AUTO",
     "Continuous autonomous monitoring — zero human babysitting.\n"
     "- Polls ClusterVersion every 15 seconds for CVO progress %\n"
     "- Tracks operator health in real-time (available/degraded/progressing)\n"
     "- Monitors node readiness as nodes reboot during upgrade\n"
     "- Live dashboard shows progress bar, operator matrix, node status\n"
     "- Updates ServiceNow CR with periodic work notes on progress",
     "30–90 min\n(cluster-paced)"),

    ("11", "Smart Completion Detection", "AI",
     "Upgrade marked complete ONLY when ALL 3 conditions met.\n"
     "- Condition 1: ClusterVersion history shows target version confirmed\n"
     "- Condition 2: ALL ClusterOperators are available and not degraded\n"
     "- Condition 3: ALL nodes are in Ready state\n"
     "- Records precise completedAt timestamp\n"
     "- Calculates and records total upgrade duration",
     "Automatic"),

    ("12", "Post-Assessment & Report Generation", "AI",
     "AI generates comprehensive post-upgrade assessment.\n"
     "- PDF Report: verified version, start/end timestamps, total duration\n"
     "- Operator health summary: total / available / degraded counts\n"
     "- Node health summary: total / ready counts\n"
     "- HTML Report with same data for web viewing\n"
     "- Both reports attached to ServiceNow CR",
     "1–2 min"),

    ("13", "ServiceNow CR Closure", "AUTO",
     "System auto-closes the Change Request with full evidence.\n"
     "- Close notes include: upgrade start time, end time, total duration\n"
     "- Operator health: 'X of Y operators healthy'\n"
     "- Node health: 'X of Y nodes ready'\n"
     "- Post-assessment PDF and HTML attached as final evidence\n"
     "- CR state set to 'Closed - Successful'\n"
     "- Complete audit trail preserved for compliance",
     "~30 sec"),
]

for i, (step, title, mode, detail, dur) in enumerate(workflow):
    r = 5 + i
    ws2.cell(r, 1, step).font = bold
    ws2.cell(r, 1).alignment = center
    ws2.cell(r, 1).border = thin

    ws2.cell(r, 2, title).font = bold
    ws2.cell(r, 2).alignment = wrap
    ws2.cell(r, 2).border = thin

    mode_cell = ws2.cell(r, 3, mode)
    mode_cell.alignment = center
    mode_cell.border = thin
    if mode == "AI":
        mode_cell.font = Font(name="Inter", bold=True, color="1E40AF", size=11)
        mode_cell.fill = light_blue
    elif mode == "AUTO":
        mode_cell.font = Font(name="Inter", bold=True, color=GREEN, size=11)
        mode_cell.fill = light_green
    elif mode == "GATE":
        mode_cell.font = Font(name="Inter", bold=True, color="9D174D", size=11)
        mode_cell.fill = PatternFill("solid", fgColor="FCE7F3")
    else:
        mode_cell.font = Font(name="Inter", bold=True, color="92400E", size=11)
        mode_cell.fill = PatternFill("solid", fgColor="FEF3C7")

    ws2.cell(r, 4, detail).font = normal
    ws2.cell(r, 4).alignment = wrap
    ws2.cell(r, 4).border = thin

    ws2.cell(r, 5, dur).font = bold_green
    ws2.cell(r, 5).alignment = center
    ws2.cell(r, 5).border = thin

    line_count = detail.count("\n") + 1
    ws2.row_dimensions[r].height = max(28, line_count * 15)

    if i % 2 == 1:
        ws2.cell(r, 1).fill = light_bg
        ws2.cell(r, 2).fill = light_bg
        ws2.cell(r, 4).fill = light_bg


# ═══════════════════════════════════════════════════════════════════
# SHEET 3: Multi-Cluster Scale
# ═══════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Multi-Cluster Scale")
col_widths(ws3, [22, 14, 24, 26, 24])

ws3.merge_cells("A1:E1")
c = ws3["A1"]
c.value = "Multi-Cluster Impact at Scale"
c.font = Font(name="Inter", bold=True, color=WHITE, size=14)
c.fill = hdr_fill
c.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 40

r = 3
headers = ["Environment Scale", "Clusters", "Manual Effort (per cycle)", "With TCS Agentic AI", "SRE Hours Reclaimed"]
hdr_row(ws3, r, 5)
for i, h in enumerate(headers, 1):
    ws3.cell(r, i, h)

scale_data = [
    ("Small Enterprise", "5", "70–200 hours\n(1–2 weeks)", "~50 min human time\n+ cluster upgrade time", "~68–198 hours saved"),
    ("Medium Enterprise", "15", "210–600 hours\n(2–4 weeks)", "~2.5 hrs human time\n+ cluster upgrade time", "~207–597 hours saved"),
    ("Large Enterprise", "50", "700–2,000 hours\n(1–2 months)", "~8 hrs human time\n+ cluster upgrade time", "~692–1,992 hours saved"),
    ("Hyperscale", "100+", "1,400–4,000 hours\n(2–4 months)", "~16 hrs human time\n+ cluster upgrade time", "~1,384–3,984 hours saved"),
]

for i, (scale, clusters, manual, auto, saved) in enumerate(scale_data):
    r = 4 + i
    ws3.cell(r, 1, scale).font = bold
    ws3.cell(r, 1).border = thin
    ws3.cell(r, 1).alignment = wrap
    ws3.cell(r, 2, clusters).font = bold
    ws3.cell(r, 2).alignment = center
    ws3.cell(r, 2).border = thin
    ws3.cell(r, 3, manual).font = bold_red
    ws3.cell(r, 3).alignment = center
    ws3.cell(r, 3).border = thin
    ws3.cell(r, 4, auto).font = bold_green
    ws3.cell(r, 4).alignment = center
    ws3.cell(r, 4).border = thin
    ws3.cell(r, 5, saved).font = bold_green
    ws3.cell(r, 5).alignment = center
    ws3.cell(r, 5).border = thin
    ws3.row_dimensions[r].height = 36
    if i % 2 == 1:
        for c2 in range(1, 6):
            ws3.cell(r, c2).fill = light_bg

# Scenario table
r = 10
ws3.merge_cells(f"A{r}:E{r}")
ws3.cell(r, 1, "REAL-WORLD UPGRADE SCENARIOS").font = bold_w
ws3.cell(r, 1).fill = accent_fill
ws3.cell(r, 1).alignment = center
for c in range(2, 6):
    ws3.cell(r, c).fill = accent_fill

r = 11
headers2 = ["Scenario", "Cluster Profile", "Manual Duration", "With TCS Agentic AI", "Outcome"]
hdr_row(ws3, r, 5)
for i, h in enumerate(headers2, 1):
    ws3.cell(r, i, h)

scenarios = [
    ("Minor Version Upgrade\n(4.14.x → 4.14.y)", "3 masters, 6 workers\nProduction", "~16 hours\n(2 days)", "~8 min human\n+ 45 min cluster", "Zero blocker misses,\nCR auto-closed with full report"),
    ("EUS-to-EUS Upgrade\n(4.12 → 4.14 via 4.13)", "3 masters, 12 workers\nProduction", "~32 hours\n(3–5 days)", "~10 min human\n+ 2.5 hrs cluster", "Multi-hop path auto-validated,\ndeprecated APIs caught"),
    ("Fleet Rolling Upgrade\n(15 clusters)", "Mixed sizes\nDev + Staging + Prod", "~300 hours\n(2–3 weeks)", "~2.5 hrs human\n+ cluster time", "Rolling orchestration,\nper-cluster reports"),
    ("Emergency CVE Patch\n(5 critical clusters)", "5 clusters\nCritical production", "~40 hours\n(expedited)", "~50 min human\n+ cluster time", "Fast-track assessment,\nexpedited CR flow"),
]

for i, (scenario, profile, manual, auto, outcome) in enumerate(scenarios):
    r = 12 + i
    ws3.cell(r, 1, scenario).font = bold
    ws3.cell(r, 1).alignment = wrap
    ws3.cell(r, 1).border = thin
    ws3.cell(r, 2, profile).font = normal
    ws3.cell(r, 2).alignment = wrap
    ws3.cell(r, 2).border = thin
    ws3.cell(r, 3, manual).font = bold_red
    ws3.cell(r, 3).alignment = center
    ws3.cell(r, 3).border = thin
    ws3.cell(r, 4, auto).font = bold_green
    ws3.cell(r, 4).alignment = center
    ws3.cell(r, 4).border = thin
    ws3.cell(r, 5, outcome).font = normal
    ws3.cell(r, 5).alignment = wrap
    ws3.cell(r, 5).border = thin
    ws3.row_dimensions[r].height = 40
    if i % 2 == 1:
        for c2 in range(1, 6):
            ws3.cell(r, c2).fill = light_bg


# Save
wb.save(OUT)
print(f"Excel saved: {OUT}")
