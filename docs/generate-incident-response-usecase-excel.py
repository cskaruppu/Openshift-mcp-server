#!/usr/bin/env python3
"""
Generate TCS Agentic AI — Use Case 2: End-to-End Incident Response
"Why is my pod crashing?" — Use Case Name, Description, and Workflow

Outputs: Excel (.xlsx) + HTML matching Use Case 1 format
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT_DIR = os.path.dirname(os.path.abspath(__file__))

wb = Workbook()

# ── Colours & Styles (matching Use Case 1) ──
PRIMARY = "1A1A2E"
ACCENT = "0F7DC2"
GREEN = "16A34A"
RED = "DC2626"
ORANGE = "E67E22"
WHITE = "FFFFFF"
LIGHT_BG = "F8FAFC"
LIGHT_GREEN = "F0FDF4"
LIGHT_RED = "FEF2F2"
LIGHT_BLUE = "EFF6FF"

hdr_font = Font(name="Inter", bold=True, color=WHITE, size=11)
hdr_fill = PatternFill("solid", fgColor=PRIMARY)
accent_fill = PatternFill("solid", fgColor=ACCENT)
light_green_fill = PatternFill("solid", fgColor="DCFCE7")
light_red_fill = PatternFill("solid", fgColor="FEE2E2")
light_blue_fill = PatternFill("solid", fgColor="DBEAFE")
light_bg_fill = PatternFill("solid", fgColor=LIGHT_BG)
bold = Font(name="Inter", bold=True, size=11)
bold_green = Font(name="Inter", bold=True, color=GREEN, size=11)
bold_red = Font(name="Inter", bold=True, color=RED, size=11)
bold_white = Font(name="Inter", bold=True, color=WHITE, size=12)
normal = Font(name="Inter", size=11)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)


def style_header_row(ws, row, cols, fill=None):
    f = fill or hdr_fill
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font
        cell.fill = f
        cell.alignment = center
        cell.border = thin_border


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════════════════════════
# SHEET 1: Executive Summary
# ═══════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Executive Summary"
set_col_widths(ws1, [35, 25, 25, 18])

# Title
ws1.merge_cells("A1:D1")
c = ws1["A1"]
c.value = "TCS Agentic AI — Use Case 2: End-to-End Incident Response"
c.font = Font(name="Inter", bold=True, color=WHITE, size=16)
c.fill = PatternFill("solid", fgColor=PRIMARY)
c.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 48

ws1.merge_cells("A2:D2")
c = ws1["A2"]
c.value = (
    '"Why is my pod crashing?" — Natural language question triggers a fully autonomous '
    "incident response pipeline: AI detects the error pattern, diagnoses root cause from "
    "8 parallel K8s API calls, auto-creates a ServiceNow incident, proposes a targeted fix "
    "with dry-run validation, applies the fix, and verifies recovery — all in under 5 minutes."
)
c.font = Font(name="Inter", size=11, color="64748B", italic=True)
c.fill = light_blue_fill
c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws1.row_dimensions[2].height = 56

# ── Use Case Identity ──
r = 4
ws1.merge_cells(f"A{r}:D{r}")
ws1.cell(r, 1, "USE CASE IDENTITY").font = bold_white
ws1.cell(r, 1).fill = accent_fill
ws1.cell(r, 1).alignment = center
for c in range(2, 5):
    ws1.cell(r, c).fill = accent_fill

fields = [
    ("Use Case Name", "End-to-End Incident Response & Application Performance Troubleshooting"),
    ("Trigger", 'User asks in AI Chat: "Why is my pod crashing?" (natural language)'),
    ("Scope", "Any pod in any namespace across any connected OpenShift cluster"),
    ("End State", "Pod healthy, ServiceNow INC resolved, before/after metrics captured, full audit trail"),
]
for i, (label, val) in enumerate(fields):
    row = 5 + i
    ws1.cell(row, 1, label).font = bold
    ws1.cell(row, 1).border = thin_border
    ws1.cell(row, 1).fill = light_blue_fill
    ws1.merge_cells(f"B{row}:D{row}")
    ws1.cell(row, 2, val).font = normal
    ws1.cell(row, 2).alignment = wrap
    ws1.cell(row, 2).border = thin_border

# ── Key Impact Numbers ──
r = 10
ws1.merge_cells(f"A{r}:D{r}")
ws1.cell(r, 1, "KEY IMPACT METRICS").font = bold_white
ws1.cell(r, 1).fill = accent_fill
ws1.cell(r, 1).alignment = center
for c in range(2, 5):
    ws1.cell(r, c).fill = accent_fill

r = 11
metrics = [
    ("Mean Time to Detect", "< 30 seconds"),
    ("Mean Time to Resolve", "< 5 minutes"),
    ("Reduction in Manual Effort", "85–90%"),
    ("Audit & Compliance Trail", "100%"),
]
style_header_row(ws1, r, 4, fill=PatternFill("solid", fgColor="0F3460"))
for i, (label, _) in enumerate(metrics, 1):
    ws1.cell(r, i, label)

r = 12
for i, (_, val) in enumerate(metrics, 1):
    cell = ws1.cell(r, i, val)
    cell.font = Font(name="Inter", bold=True, color=GREEN, size=14)
    cell.alignment = center
    cell.fill = light_green_fill
    cell.border = thin_border
ws1.row_dimensions[12].height = 36

# ── The Challenge ──
r = 14
ws1.merge_cells(f"A{r}:D{r}")
ws1.cell(r, 1, "THE CHALLENGE: MANUAL INCIDENT RESPONSE").font = bold_white
ws1.cell(r, 1).fill = PatternFill("solid", fgColor=RED)
ws1.cell(r, 1).alignment = center
for c in range(2, 5):
    ws1.cell(r, c).fill = PatternFill("solid", fgColor=RED)

challenges = [
    ("Slow Detection", "SREs rely on alerts that may fire minutes after a pod crash — OOMKilled and CrashLoopBackOff often go unnoticed until end-users report impact"),
    ("Manual Root Cause Analysis", "SREs manually run kubectl describe, kubectl logs, kubectl get events across pods and namespaces — often taking 30–60 minutes per incident"),
    ("Knowledge Silos", "Troubleshooting expertise lives in senior SREs' heads; junior team members struggle with unfamiliar error patterns, extending resolution time"),
    ("ServiceNow Gaps", "Incident tickets created manually, often incomplete, missing root cause details, no correlation between detection and resolution evidence"),
    ("No Recovery Validation", "After applying a fix, SREs manually check pod status, restart counts, and resource metrics — no automated proof that the fix actually worked"),
]
r = 15
ws1.cell(r, 1, "Pain Point").font = bold
ws1.cell(r, 2, "Description").font = bold
ws1.merge_cells(f"B{r}:D{r}")
for c in range(1, 5):
    ws1.cell(r, c).fill = light_red_fill
    ws1.cell(r, c).border = thin_border

for i, (pain, desc) in enumerate(challenges):
    row = 16 + i
    ws1.cell(row, 1, pain).font = Font(name="Inter", bold=True, color=RED, size=11)
    ws1.cell(row, 1).border = thin_border
    ws1.merge_cells(f"B{row}:D{row}")
    ws1.cell(row, 2, desc).font = normal
    ws1.cell(row, 2).alignment = wrap
    ws1.cell(row, 2).border = thin_border

# ── The Solution ──
r = 22
ws1.merge_cells(f"A{r}:D{r}")
ws1.cell(r, 1, 'THE SOLUTION: "WHY IS MY POD CRASHING?" END-TO-END PIPELINE').font = bold_white
ws1.cell(r, 1).fill = PatternFill("solid", fgColor=GREEN)
ws1.cell(r, 1).alignment = center
for c in range(2, 5):
    ws1.cell(r, c).fill = PatternFill("solid", fgColor=GREEN)

capabilities = [
    ("Natural Language Understanding", 'NLU normalizes "pod crashing" to incident_response intent with 0.88 confidence — routes to dedicated pipeline'),
    ("8 Parallel K8s API Calls", "Simultaneously fetches pod describe, logs, previous logs, events, metrics, node conditions, deployment config, HPA status"),
    ("12-Pattern Error Engine", "Detects OOMKilled, CrashLoopBackOff, ImagePullBackOff, CreateContainerConfigError, Evicted, Pending, Probe failures, and more"),
    ("AI Root Cause Analysis", "LLM analyzes all collected evidence, identifies root cause, calculates blast radius, and determines affected services"),
    ("Smart Memory Calculation", "For OOMKilled: calculates optimal memory limit based on restart count and usage — rounds to standard values (256Mi, 512Mi, 1Gi, etc.)"),
    ("ServiceNow Auto-Ticketing", "Creates INC with severity mapping (CRITICAL→SEV-2, WARNING→SEV-3), root cause, evidence, and fix proposals — zero manual input"),
    ("Fix Proposals with Dry Run", "Generates targeted oc/kubectl commands with risk scoring; dry-run validates before real execution"),
    ("Recovery Verification", "Post-fix: polls deployment readiness every 3s (60s timeout), verifies all pods healthy, captures before/after metrics, auto-closes INC"),
]

r = 23
ws1.cell(r, 1, "Capability").font = bold
ws1.cell(r, 2, "Description").font = bold
ws1.merge_cells(f"B{r}:D{r}")
for c in range(1, 5):
    ws1.cell(r, c).fill = light_green_fill
    ws1.cell(r, c).border = thin_border

for i, (cap, desc) in enumerate(capabilities):
    row = 24 + i
    ws1.cell(row, 1, cap).font = Font(name="Inter", bold=True, color=GREEN, size=11)
    ws1.cell(row, 1).border = thin_border
    ws1.cell(row, 1).alignment = wrap
    ws1.merge_cells(f"B{row}:D{row}")
    ws1.cell(row, 2, desc).font = normal
    ws1.cell(row, 2).alignment = wrap
    ws1.cell(row, 2).border = thin_border


# ═══════════════════════════════════════════════════════════════════
# SHEET 2: Use Case Workflow (Stage-by-Stage)
# ═══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Workflow — Stage by Stage")
set_col_widths(ws2, [6, 24, 10, 24, 40])

ws2.merge_cells("A1:E1")
c = ws2["A1"]
c.value = 'Use Case 2 Workflow: "Why is my pod crashing?" — What Happens at Each Stage'
c.font = Font(name="Inter", bold=True, color=WHITE, size=14)
c.fill = hdr_fill
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 40

ws2.merge_cells("A2:E2")
c = ws2["A2"]
c.value = "User types a natural language question in AI Chat. The system executes 10 stages autonomously — from NLU parsing to verified recovery — with ServiceNow integration and full audit trail."
c.font = Font(name="Inter", size=10, color="64748B", italic=True)
c.fill = light_blue_fill
c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws2.row_dimensions[2].height = 36

r = 4
headers = ["Stage", "Step Name", "Duration", "What Happens", "Technical Detail"]
style_header_row(ws2, r, 5)
for i, h in enumerate(headers, 1):
    ws2.cell(r, i, h)

stages = [
    (
        "01",
        "Natural Language\nQuery Parsing",
        "~1 sec",
        'User asks "Why is my pod crashing?" — NLU engine parses the question',
        'normalizeCompounds() matches "pod crash*" → replaces with incident_response pseudo-resource. '
        "tokenize() splits into tokens. VERB_TABLE maps \"why\" → diagnose intent. "
        "Resource detector finds incident_response → returns {intent: incident_response, confidence: 0.88}. "
        "Bypasses cache (CACHE_BYPASS_INTENTS) — always hits live cluster data.",
    ),
    (
        "02",
        "Cluster Context\nGathering",
        "~3 sec",
        "8 parallel K8s API calls fetch all evidence about the crashing pod",
        "If pod name detected: (1) GET pods/{name}, (2) GET events?involvedObject={name}, "
        "(3) GET pods/{name}/log?tailLines=80, (4) GET pods/{name}/log?previous=true, "
        "(5) GET metrics.k8s.io pods/{name}. "
        "If no pod name: GET /api/v1/pods (all), GET /api/v1/nodes, "
        "GET /api/v1/events?type=Warning, GET clusterversions, GET clusteroperators — "
        "then filters to problemPods (CrashLoopBackOff, OOMKilled, ImagePullBackOff, etc.).",
    ),
    (
        "03",
        "Pod Doctor\nDiagnosis",
        "~2 sec",
        "12-pattern error detection engine analyzes the pod state",
        "diagnosePod() counts totalRestarts, identifies crashingContainers via regex "
        "(CrashLoopBackOff|Error|OOMKilled|Terminated). Per-container diagnoseContainer() "
        "checks in priority order: (1) OOMKilled — exitCode 137 or terminationReason, "
        "(2) CrashLoopBackOff — restarts > 5 + exitCodeMeaning(), "
        "(3) ImagePullBackOff, (4) CreateContainerConfigError, (5) Evicted, "
        "(6) ReadinessProbe, (7) LivenessProbe, (8) InitContainer failure. "
        "For CrashLoop: detectConfigIssue() checks 8 sub-patterns (ConnectionRefused, "
        "MissingEnvVar, PermissionDenied, MissingFile, BadFlag, ConfigParseError, PortConflict).",
    ),
    (
        "04",
        "Severity\nAssessment",
        "~1 sec",
        "Classifies incident as CRITICAL / WARNING / INFO based on evidence",
        "OOMKilled → CRITICAL. CrashLoopBackOff (restarts > 5) → CRITICAL. "
        "ImagePullBackOff → WARNING. Evicted → CRITICAL. "
        "ReadinessProbe → WARNING. LivenessProbe → CRITICAL. "
        "Severity drives ServiceNow urgency mapping and fix priority ordering.",
    ),
    (
        "05",
        "Smart Fix\nProposal Generation",
        "~2 sec",
        "AI generates targeted fix commands with risk scoring",
        "For OOMKilled: calculateSmartMemoryLimit() computes optimal memory — "
        "multiplier scales with restarts: 2x (≤2), 3x (≤5), 4x (≤10), 5x (>10). "
        "Considers 2.5x current usage if higher. Rounds to nice values "
        "(128Mi, 256Mi, 384Mi, 512Mi, 768Mi, 1Gi, 1.5Gi, 2Gi, 3Gi, 4Gi, 6Gi, 8Gi). "
        "resolveDeploymentResource() finds owning Deployment/StatefulSet. "
        "Generates: oc set resources deployment/<name> --limits=memory=<calculated>. "
        "Each fix includes: title, action, command, risk score, rollback instructions.",
    ),
    (
        "06",
        "ServiceNow\nINC Creation",
        "~3 sec",
        "Auto-creates incident ticket with severity, root cause, evidence, and fixes",
        "createServiceNowIncident() called with: short_description = "
        '"[SEV-2] OOMKilled — <deploy> — Auto-generated by TCS Agentic AI", '
        "description = severity + pod + namespace + ROOT CAUSE ANALYSIS + EVIDENCE + FIX, "
        "urgency = 2 (CRITICAL→2, WARNING→3), impact = 2. "
        "Returns incidentSysId + incidentNumber (e.g., INC0012345). "
        "Appends notice to chat: \"ServiceNow Incident INC0012345 auto-created for SEV-2\". "
        "Incident timeline built from pod events + AI detection event + fix-ready event.",
    ),
    (
        "07",
        "Fix Card Displayed\nin AI Chat",
        "Instant",
        "Interactive fix proposal card rendered with Apply Fix and Dry Run buttons",
        "appendFixProposals() injects @@FIX_PROPOSAL|{json}@@ token into reply. "
        "Frontend ChatTokens.jsx parseSegments() detects the token, renders ActionCard "
        "with: diagnosis summary, root cause, evidence list, severity badge, "
        "incident timeline (collapsible), and per-fix buttons (Apply Fix / Dry Run). "
        "Each fix card shows: command, risk level, description, rollback instructions.",
    ),
    (
        "08",
        "Dry Run\nValidation",
        "~5 sec",
        "User clicks Dry Run — fix command executed with --dry-run flag",
        "POST /api/alerts/execute-fix with {command, dryRun: true}. "
        "Rate-limited (burst=5). Guardrails preflightCheck() classifies command. "
        "executeFixCommand() runs with ?dryRun=All — K8s API validates the patch "
        "without applying. Blocks dangerous patterns (delete namespace/crd/node, "
        "cordon/drain, exec rm). Returns validation result — user sees confirmation.",
    ),
    (
        "09",
        "Apply Fix &\nBefore/After Metrics",
        "~30 sec",
        "User clicks Apply Fix — command executed, before/after metrics captured",
        "POST /api/alerts/execute-fix with {command, dryRun: false, captureMetrics: true, "
        "incidentSysId, incidentNumber}. BEFORE: fetches pods by label selector, "
        "captures memoryLimit, memoryUsage, cpuUsage, restarts, status. "
        "EXECUTE: e.g., PATCH deployment to set new memory limits. "
        "Audit logged (userId, command, classification, duration). "
        "Learning loop records fix for future similar-incident retrieval.",
    ),
    (
        "10",
        "Recovery Verification\n& INC Auto-Close",
        "~30 sec",
        "Polls deployment readiness, verifies pods healthy, auto-closes ServiceNow INC",
        "Rollout poll every 3s (60s timeout): checks readyReplicas ≥ desired, "
        "updatedReplicas ≥ desired, unavailableReplicas = 0. "
        "Pod health: fetches all pods by label, checks phase=Running + ready=true. "
        "AFTER metrics: re-captures memoryLimit, memoryUsage, restarts, status. "
        "On success: snowResolveIncident() closes INC with closeCode='Solved (Permanently)', "
        "workNotes include fix command, namespace, severity, diagnosis, validation results. "
        "UI shows RecoveryTimeline (step-by-step) + BeforeAfterMetrics (red/green comparison).",
    ),
]

for i, (num, step, dur, what, detail) in enumerate(stages):
    row = 5 + i
    ws2.cell(row, 1, num).font = bold
    ws2.cell(row, 1).alignment = center
    ws2.cell(row, 1).border = thin_border
    ws2.cell(row, 2, step).font = bold
    ws2.cell(row, 2).alignment = wrap
    ws2.cell(row, 2).border = thin_border
    ws2.cell(row, 3, dur).font = bold_green
    ws2.cell(row, 3).alignment = center
    ws2.cell(row, 3).border = thin_border
    ws2.cell(row, 4, what).font = normal
    ws2.cell(row, 4).alignment = wrap
    ws2.cell(row, 4).border = thin_border
    ws2.cell(row, 5, detail).font = Font(name="Inter", size=10, color="475569")
    ws2.cell(row, 5).alignment = wrap
    ws2.cell(row, 5).border = thin_border
    ws2.row_dimensions[row].height = 80
    if i % 2 == 1:
        for c in range(1, 6):
            ws2.cell(row, c).fill = light_bg_fill

# Total row
row = 15
ws2.merge_cells(f"A{row}:B{row}")
ws2.cell(row, 1, "TOTAL END-TO-END").font = Font(name="Inter", bold=True, color=WHITE, size=12)
ws2.cell(row, 1).fill = PatternFill("solid", fgColor=PRIMARY)
ws2.cell(row, 1).alignment = center
ws2.cell(row, 1).border = thin_border
ws2.cell(row, 2).fill = PatternFill("solid", fgColor=PRIMARY)
ws2.cell(row, 2).border = thin_border
ws2.cell(row, 3, "< 5 min").font = Font(name="Inter", bold=True, color=WHITE, size=14)
ws2.cell(row, 3).fill = PatternFill("solid", fgColor=GREEN)
ws2.cell(row, 3).alignment = center
ws2.cell(row, 3).border = thin_border
ws2.merge_cells(f"D{row}:E{row}")
ws2.cell(row, 4, "From natural language question to verified recovery with ServiceNow INC closed").font = Font(name="Inter", bold=True, color=WHITE, size=11)
ws2.cell(row, 4).fill = PatternFill("solid", fgColor=PRIMARY)
ws2.cell(row, 4).alignment = center
ws2.cell(row, 4).border = thin_border
ws2.cell(row, 5).fill = PatternFill("solid", fgColor=PRIMARY)
ws2.cell(row, 5).border = thin_border
ws2.row_dimensions[row].height = 32


# ═══════════════════════════════════════════════════════════════════
# SHEET 3: Manual vs Automated
# ═══════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Manual vs Automated")
set_col_widths(ws3, [28, 22, 22, 14, 30])

ws3.merge_cells("A1:E1")
c = ws3["A1"]
c.value = "Side-by-Side Comparison: Manual SRE vs. TCS Agentic AI"
c.font = Font(name="Inter", bold=True, color=WHITE, size=14)
c.fill = hdr_fill
c.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 40

r = 3
headers = ["Incident Phase", "Manual (SRE Time)", "Automated (AI Time)", "Time Saved", "How It Works"]
style_header_row(ws3, r, 5)
for i, h in enumerate(headers, 1):
    ws3.cell(r, i, h)

comparison = [
    ("Detect & Triage", "5–15 min (wait for alert)", "< 30 sec (NLU parse)", "95%", "NLU normalizes 'pod crash' → incident_response intent at 0.88 confidence"),
    ("Collect Evidence", "15–30 min (kubectl describe,\nlogs, events, metrics)", "~3 sec (8 parallel API calls)", "98%", "Parallel: pod describe + logs + prev logs + events + metrics + node + deploy + HPA"),
    ("Identify Error Pattern", "10–20 min (read logs,\ncorrelate events)", "~2 sec (12-pattern engine)", "98%", "diagnosePod() matches OOMKilled/CrashLoop/ImagePull/Evicted/Probe + 7 more"),
    ("Root Cause Analysis", "30–60 min (senior SRE\nexpertise required)", "~5 sec (LLM analysis)", "97%", "AI cross-references all evidence, identifies root cause, calculates blast radius"),
    ("Severity Classification", "5–10 min", "~1 sec", "97%", "Rule-based: OOMKilled→CRITICAL, ImagePull→WARNING — drives INC urgency"),
    ("Create ServiceNow INC", "15–30 min (manual form)", "~3 sec (API auto-create)", "97%", "createServiceNowIncident() with severity, root cause, evidence, fix proposals"),
    ("Develop Fix", "20–45 min (research +\ncraft command)", "~2 sec (auto-generated)", "97%", "Smart memory calc for OOM; rollback for CrashLoop; credential fix for ImagePull"),
    ("Validate Fix (Dry Run)", "15–30 min (test env)", "~5 sec (K8s dry-run API)", "96%", "executeFixCommand() with ?dryRun=All — validates patch without applying"),
    ("Apply Fix", "10–20 min (apply + monitor)", "~30 sec (one-click apply)", "95%", "PATCH deployment via K8s API + before-metrics captured automatically"),
    ("Verify Recovery", "15–30 min (manual checks)", "~30 sec (auto-poll)", "96%", "Poll every 3s: readyReplicas ≥ desired + allPods Running+Ready + after-metrics"),
    ("Document & Close INC", "20–40 min", "~5 sec (auto-close)", "98%", "snowResolveIncident() with fix details, validation results, before/after metrics"),
]

for i, (phase, manual, auto, saved, how) in enumerate(comparison):
    row = 4 + i
    ws3.cell(row, 1, phase).font = bold
    ws3.cell(row, 1).border = thin_border
    ws3.cell(row, 1).alignment = wrap
    ws3.cell(row, 2, manual).font = Font(name="Inter", color=RED, size=11)
    ws3.cell(row, 2).alignment = center
    ws3.cell(row, 2).border = thin_border
    ws3.cell(row, 3, auto).font = Font(name="Inter", color=GREEN, size=11)
    ws3.cell(row, 3).alignment = center
    ws3.cell(row, 3).border = thin_border
    ws3.cell(row, 4, saved).font = bold_green
    ws3.cell(row, 4).alignment = center
    ws3.cell(row, 4).border = thin_border
    ws3.cell(row, 5, how).font = Font(name="Inter", size=10, color="475569")
    ws3.cell(row, 5).alignment = wrap
    ws3.cell(row, 5).border = thin_border
    if i % 2 == 1:
        for c in range(1, 6):
            ws3.cell(row, c).fill = light_bg_fill

# Total
row = 15
ws3.cell(row, 1, "TOTAL (per incident)").font = Font(name="Inter", bold=True, color=WHITE, size=12)
ws3.cell(row, 1).fill = PatternFill("solid", fgColor=PRIMARY)
ws3.cell(row, 1).border = thin_border
ws3.cell(row, 2, "2.5–5.5 hours").font = Font(name="Inter", bold=True, color=WHITE, size=12)
ws3.cell(row, 2).fill = PatternFill("solid", fgColor=RED)
ws3.cell(row, 2).alignment = center
ws3.cell(row, 2).border = thin_border
ws3.cell(row, 3, "< 5 minutes").font = Font(name="Inter", bold=True, color=WHITE, size=12)
ws3.cell(row, 3).fill = PatternFill("solid", fgColor=GREEN)
ws3.cell(row, 3).alignment = center
ws3.cell(row, 3).border = thin_border
ws3.cell(row, 4, "~97%").font = Font(name="Inter", bold=True, color=WHITE, size=12)
ws3.cell(row, 4).fill = PatternFill("solid", fgColor=GREEN)
ws3.cell(row, 4).alignment = center
ws3.cell(row, 4).border = thin_border
ws3.cell(row, 5, "End-to-End AI Orchestration").font = Font(name="Inter", bold=True, color=WHITE, size=11)
ws3.cell(row, 5).fill = PatternFill("solid", fgColor=PRIMARY)
ws3.cell(row, 5).alignment = center
ws3.cell(row, 5).border = thin_border
ws3.row_dimensions[row].height = 32


# ═══════════════════════════════════════════════════════════════════
# SHEET 4: Error Patterns & Detection
# ═══════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Error Patterns & Detection")
set_col_widths(ws4, [22, 36, 14, 34])

ws4.merge_cells("A1:D1")
c = ws4["A1"]
c.value = "12 Error Patterns Detected by Pod Doctor Engine"
c.font = Font(name="Inter", bold=True, color=WHITE, size=14)
c.fill = hdr_fill
c.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 40

r = 3
headers = ["Error Pattern", "How Pod Doctor Detects It", "Severity", "Auto-Generated Fix"]
style_header_row(ws4, r, 4)
for i, h in enumerate(headers, 1):
    ws4.cell(r, i, h)

patterns = [
    ("OOMKilled", "exitCode === 137 OR terminationReason === 'OOMKilled'; memory usage vs limits analysis", "CRITICAL",
     "oc set resources deployment/<name> --limits=memory=<calculated>\ncalculateSmartMemoryLimit(): 2x–5x based on restarts, rounded to 128Mi–8Gi"),
    ("CrashLoopBackOff", "container.state === 'CrashLoopBackOff' OR restarts > 5 && exitCode !== 0; exitCodeMeaning() lookup", "CRITICAL",
     "Rollback to last stable image OR fix detected config issue\ndetectConfigIssue() checks 8 sub-patterns"),
    ("ImagePullBackOff", "container.waiting.reason === 'ImagePullBackOff'; registry + image name analysis", "WARNING",
     "Verify image tag exists, recreate pull secret\noc create secret docker-registry ..."),
    ("CreateContainerConfigError", "container.waiting.reason === 'CreateContainerConfigError'; missing ConfigMap/Secret detection", "WARNING",
     "Identify missing ConfigMap/Secret, recreate from source\noc create configmap <name> --from-file=..."),
    ("Evicted", "pod.status.phase === 'Failed', reason === 'Evicted'; node resource pressure analysis", "CRITICAL",
     "Set ephemeral storage limits, clean disk, rebalance workloads\noc set resources ... --limits=ephemeral-storage=..."),
    ("Pending (Unschedulable)", "pod.phase === 'Pending'; insufficient CPU/memory or node affinity failures", "WARNING",
     "Reduce resource requests OR suggest node scaling\noc scale machinesets ... --replicas=+1"),
    ("ReadinessProbe Failure", "container not ready; events with reason 'Unhealthy' + /readiness/", "WARNING",
     "Adjust probe thresholds: timeoutSeconds, failureThreshold\noc patch deployment ... -p '{...}'"),
    ("LivenessProbe Failure", "container restarting; events with reason 'Unhealthy' + /liveness/", "CRITICAL",
     "Increase initialDelaySeconds, adjust probe config\noc patch deployment ... -p '{...}'"),
    ("Init Container Failure", "initContainer in CrashLoopBackOff or Error; dependency chain analysis", "WARNING",
     "Fix init script or service dependency\nCheck init container logs for root cause"),
    ("Volume Mount Error", "container waiting on volume; PVC pending or mount failure detected", "WARNING",
     "Check PVC status, fix StorageClass, verify zone affinity\noc get pvc -n <ns> -o wide"),
    ("Node NotReady Impact", "pod on NotReady node; node condition analysis + cordon status", "CRITICAL",
     "Migrate affected pods, analyze node conditions\noc adm drain <node> --ignore-daemonsets"),
    ("Resource Quota Exceeded", "pod rejected by admission; namespace quota analysis", "WARNING",
     "Increase quota or reduce resource requests\noc patch resourcequota ... -p '{...}'"),
]

for i, (pattern, method, sev, fix) in enumerate(patterns):
    row = 4 + i
    ws4.cell(row, 1, pattern).font = bold
    ws4.cell(row, 1).border = thin_border
    ws4.cell(row, 1).alignment = wrap
    ws4.cell(row, 2, method).font = normal
    ws4.cell(row, 2).alignment = wrap
    ws4.cell(row, 2).border = thin_border
    cell_sev = ws4.cell(row, 3, sev)
    cell_sev.alignment = center
    cell_sev.border = thin_border
    if sev == "CRITICAL":
        cell_sev.font = Font(name="Inter", bold=True, color=RED, size=10)
        cell_sev.fill = light_red_fill
    else:
        cell_sev.font = Font(name="Inter", bold=True, color=ORANGE, size=10)
        cell_sev.fill = PatternFill("solid", fgColor="FEF3C7")
    ws4.cell(row, 4, fix).font = Font(name="Inter", size=10, color="475569")
    ws4.cell(row, 4).alignment = wrap
    ws4.cell(row, 4).border = thin_border
    ws4.row_dimensions[row].height = 48
    if i % 2 == 1:
        ws4.cell(row, 1).fill = light_bg_fill
        ws4.cell(row, 2).fill = light_bg_fill
        ws4.cell(row, 4).fill = light_bg_fill


# ═══════════════════════════════════════════════════════════════════
# SHEET 5: Real-World Scenarios
# ═══════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Real-World Scenarios")
set_col_widths(ws5, [20, 22, 18, 18, 36])

ws5.merge_cells("A1:E1")
c = ws5["A1"]
c.value = "Real-World Scenarios: What Happens When You Ask \"Why is my pod crashing?\""
c.font = Font(name="Inter", bold=True, color=WHITE, size=14)
c.fill = hdr_fill
c.alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[1].height = 40

r = 3
headers = ["Scenario", "Root Cause Found", "Manual Resolution", "With TCS Agentic AI", "What the User Sees"]
style_header_row(ws5, r, 5)
for i, h in enumerate(headers, 1):
    ws5.cell(r, i, h)

scenarios = [
    (
        "OOMKilled\nJava memory leak",
        "Container limit 256Mi,\nactual usage 280Mi+,\nexitCode 137",
        "~2 hours\nSSH → describe → logs →\nidentify OOM → patch",
        "~3 min\nDetect → RCA → INC →\npatch to 512Mi → verify",
        "AI Chat shows: diagnosis card with OOMKilled root cause, memory usage chart, "
        "INC0012345 auto-created, 'Apply Fix' button patches to 512Mi, "
        "RecoveryTimeline shows pod restarting → healthy, BeforeAfterMetrics shows 256Mi→512Mi",
    ),
    (
        "CrashLoopBackOff\nMissing ConfigMap",
        "Deployment references\ndeleted ConfigMap 'app-config',\nexit code 1",
        "~1.5 hours\nCheck events → logs →\ntrace config → recreate",
        "~4 min\nDetect → missing config →\nINC → recreate → verify",
        "AI Chat shows: CrashLoopBackOff diagnosis, detectConfigIssue() found 'MissingFile', "
        "evidence includes log lines, INC auto-created with dependency chain, "
        "fix proposes ConfigMap recreation",
    ),
    (
        "ImagePullBackOff\nExpired pull secret",
        "Pull secret expired,\nregistry returns 401\nfor private image",
        "~1 hour\nCheck events → test →\nrenew secret → rollout",
        "~2 min\nDetect → auth failure →\nINC → renew → verify",
        "AI Chat shows: ImagePullBackOff diagnosis, registry URL identified, "
        "INC auto-created as SEV-3 (WARNING), fix proposes new pull secret creation",
    ),
    (
        "LivenessProbe\nSlow startup",
        "App takes 45s to start,\nliveness probe fails at 30s,\npod keeps restarting",
        "~1.5 hours\nCorrelate restarts with\nprobe config → adjust",
        "~3 min\nDetect → probe analysis →\nINC → patch probe → verify",
        "AI Chat shows: LivenessProbe CRITICAL diagnosis, events show 'Unhealthy' pattern, "
        "fix proposes initialDelaySeconds increase from 30→60, INC auto-created",
    ),
    (
        "Evicted Pods\nDisk pressure",
        "Ephemeral storage full,\nlogs filling /var/log,\n5 pods evicted",
        "~2 hours\nSSH to node → check →\nclean → reschedule",
        "~4 min\nDetect → pressure found →\nINC → limits + clean →\nreschedule → verify",
        "AI Chat shows: Evicted CRITICAL diagnosis, node disk pressure evidence, "
        "INC created for all 5 pods, fix proposes ephemeral-storage limits, "
        "recovery confirms all 5 pods rescheduled",
    ),
    (
        "Node NotReady\n12 pods displaced",
        "Worker node kernel panic,\n12 pods across\n4 namespaces affected",
        "~3 hours\nIdentify node → list →\nmigrate → verify each",
        "~5 min\nBlast radius mapped →\nINC → migrate →\nverify all 12 pods",
        "AI Chat shows: blast radius analysis across 4 namespaces, "
        "single INC with full impact list, pod migration triggered, "
        "RecoveryTimeline tracks all 12 pods returning to healthy",
    ),
]

for i, (scenario, cause, manual, auto, user_sees) in enumerate(scenarios):
    row = 4 + i
    ws5.cell(row, 1, scenario).font = bold
    ws5.cell(row, 1).alignment = wrap
    ws5.cell(row, 1).border = thin_border
    ws5.cell(row, 2, cause).font = normal
    ws5.cell(row, 2).alignment = wrap
    ws5.cell(row, 2).border = thin_border
    ws5.cell(row, 3, manual).font = bold_red
    ws5.cell(row, 3).alignment = center
    ws5.cell(row, 3).border = thin_border
    ws5.cell(row, 4, auto).font = bold_green
    ws5.cell(row, 4).alignment = center
    ws5.cell(row, 4).border = thin_border
    ws5.cell(row, 5, user_sees).font = Font(name="Inter", size=10, color="475569")
    ws5.cell(row, 5).alignment = wrap
    ws5.cell(row, 5).border = thin_border
    ws5.row_dimensions[row].height = 72
    if i % 2 == 1:
        for c in range(1, 6):
            ws5.cell(row, c).fill = light_bg_fill


# ═══════════════════════════════════════════════════════════════════
# SHEET 6: Safety & Key Benefits
# ═══════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Safety & Key Benefits")
set_col_widths(ws6, [28, 50])

ws6.merge_cells("A1:B1")
c = ws6["A1"]
c.value = "Built-in Safety Guardrails"
c.font = Font(name="Inter", bold=True, color=WHITE, size=14)
c.fill = hdr_fill
c.alignment = Alignment(horizontal="center", vertical="center")
ws6.row_dimensions[1].height = 40

r = 3
style_header_row(ws6, r, 2)
ws6.cell(r, 1, "Guardrail")
ws6.cell(r, 2, "How It Works")

guardrails = [
    ("Dry Run Before Apply", "Every fix can be validated with K8s ?dryRun=All before real execution — user sees confirmation before committing"),
    ("Dangerous Command Blocking", "fix-executor.js blocks: delete namespace/crd/node, cordon/drain, exec rm, delete all — hard-coded deny list"),
    ("Rate Limiting", "execute-fix endpoint rate-limited at burst=5, refill=0.1/sec — prevents automated mass changes"),
    ("User Confirmation Required", "Every fix requires explicit button click — no auto-apply; fix card shows command, risk level, and rollback instructions"),
    ("Guardrails Preflight", "preflightCheck() classifies every command before execution — blocks high-risk operations if guardrails pillar enabled"),
    ("StatefulSet Protection", "Pod doctor detects StatefulSets and prevents force-delete; flags data-loss risks explicitly in fix proposals"),
    ("Recovery Timeout", "60-second timeout on rollout polling — if pod fails to recover, INC remains open for manual review, user notified"),
    ("Complete Audit Trail", "Every fix execution logged: userId, command, classification, success/failure, duration, before/after metrics"),
]

for i, (guard, desc) in enumerate(guardrails):
    row = 4 + i
    ws6.cell(row, 1, guard).font = Font(name="Inter", bold=True, color=RED, size=11)
    ws6.cell(row, 1).border = thin_border
    ws6.cell(row, 1).alignment = wrap
    ws6.cell(row, 2, desc).font = normal
    ws6.cell(row, 2).alignment = wrap
    ws6.cell(row, 2).border = thin_border
    if i % 2 == 1:
        for c in range(1, 3):
            ws6.cell(row, c).fill = light_bg_fill

# Key Benefits section
r = 13
ws6.merge_cells(f"A{r}:B{r}")
ws6.cell(r, 1, "KEY BENEFITS").font = bold_white
ws6.cell(r, 1).fill = PatternFill("solid", fgColor=GREEN)
ws6.cell(r, 1).alignment = center
ws6.cell(r, 2).fill = PatternFill("solid", fgColor=GREEN)

r = 14
style_header_row(ws6, r, 2)
ws6.cell(r, 1, "Benefit")
ws6.cell(r, 2, "Description")

benefits = [
    ("97% Time Reduction", "From 2.5–5.5 hours manual SRE effort to under 5 minutes — detection through verified recovery with ServiceNow INC closed"),
    ("Natural Language Interface", 'Ask "why is my pod crashing?" in plain English — no kubectl expertise needed; junior SREs get senior-level diagnosis'),
    ("12-Pattern Intelligence", "Comprehensive error detection: OOMKilled, CrashLoopBackOff, ImagePullBackOff, Evicted, Probe failures, and 7 more — with targeted fix for each"),
    ("ServiceNow Native ITSM", "Auto-create INC with severity mapping, root cause, evidence, timeline — auto-close with resolution proof when fix verified"),
    ("Before/After Proof", "Quantified evidence: memory 256Mi→512Mi, restarts 47→0, status CrashLoopBackOff→Running — shows fix effectiveness"),
    ("Smart Fix Generation", "OOMKilled gets calculated memory limits (2x–5x based on restarts); CrashLoop gets config issue detection (8 sub-patterns)"),
    ("Recovery Verification", "Auto-poll every 3s: replicas ready + all pods Running+Ready + metrics normalized — not just 'command succeeded'"),
    ("Knowledge Democratization", "AI captures troubleshooting expertise — every SRE gets consistent, expert-level diagnosis regardless of experience"),
]

for i, (benefit, desc) in enumerate(benefits):
    row = 15 + i
    ws6.cell(row, 1, benefit).font = bold_green
    ws6.cell(row, 1).border = thin_border
    ws6.cell(row, 1).alignment = wrap
    ws6.cell(row, 2, desc).font = normal
    ws6.cell(row, 2).alignment = wrap
    ws6.cell(row, 2).border = thin_border
    if i % 2 == 1:
        for c in range(1, 3):
            ws6.cell(row, c).fill = light_bg_fill


# ── Save Excel ──
excel_path = os.path.join(OUT_DIR, "TCS-Agentic-AI-Incident-Response-UseCase.xlsx")
wb.save(excel_path)
print(f"Excel saved: {excel_path}")


# ═══════════════════════════════════════════════════════════════════
# HTML GENERATION
# ═══════════════════════════════════════════════════════════════════

def generate_html():
    # Lifecycle steps
    lifecycle_steps = [
        ("01", "NL Query Parsing", "NLU normalizes 'pod crash' → incident_response at 0.88 confidence"),
        ("02", "8 Parallel K8s API Calls", "Pod describe + logs + prev logs + events + metrics + node + deploy + HPA"),
        ("03", "Pod Doctor Diagnosis", "12-pattern engine: OOMKilled, CrashLoop, ImagePull, Evicted + 8 more"),
        ("04", "Severity Assessment", "CRITICAL / WARNING / INFO based on error pattern + restart count"),
        ("05", "Smart Fix Generation", "Targeted commands with risk score; smart memory calc for OOMKilled"),
        ("06", "ServiceNow INC Creation", "Auto-create with severity mapping, root cause, evidence, fixes"),
        ("07", "Fix Card in AI Chat", "Interactive card: diagnosis, evidence, Apply Fix / Dry Run buttons"),
        ("08", "Dry Run Validation", "K8s ?dryRun=All validates patch without applying; blocks dangerous ops"),
        ("09", "Apply Fix + Metrics", "Execute command, capture before/after memory, restarts, status"),
        ("10", "Recovery + INC Close", "Poll 3s/60s timeout; verify pods healthy; auto-close ServiceNow INC"),
    ]
    lifecycle_html = ""
    for num, name, desc in lifecycle_steps:
        lifecycle_html += f'''
        <div class="wf-step">
          <div class="wf-num">{num}</div>
          <div class="wf-name">{name}</div>
          <div class="wf-desc">{desc}</div>
        </div>'''

    # Comparison rows
    comp_data = [
        ("Detect & Triage", "5–15 min", "< 30 sec", "95%"),
        ("Collect Evidence", "15–30 min", "~3 sec", "98%"),
        ("Identify Error Pattern", "10–20 min", "~2 sec", "98%"),
        ("Root Cause Analysis", "30–60 min", "~5 sec", "97%"),
        ("Create ServiceNow INC", "15–30 min", "~3 sec", "97%"),
        ("Develop Fix", "20–45 min", "~2 sec", "97%"),
        ("Dry Run + Apply Fix", "25–50 min", "~35 sec", "96%"),
        ("Verify + Close INC", "35–70 min", "~35 sec", "97%"),
    ]
    comp_rows = ""
    for phase, manual, auto, saved in comp_data:
        comp_rows += f'''
        <tr>
          <td class="phase">{phase}</td>
          <td class="manual">{manual}</td>
          <td class="auto">{auto}</td>
          <td class="saved">{saved}</td>
        </tr>'''

    # Error patterns
    pattern_data = [
        ("OOMKilled", "CRITICAL", "exitCode 137 / terminationReason", "Smart memory limit: 2x–5x restarts"),
        ("CrashLoopBackOff", "CRITICAL", "restarts > 5, exitCode !== 0", "Rollback image / fix config (8 sub-patterns)"),
        ("ImagePullBackOff", "WARNING", "waiting.reason = ImagePullBackOff", "Verify image, recreate pull secret"),
        ("CreateContainerConfigError", "WARNING", "Missing ConfigMap/Secret", "Identify + recreate missing config"),
        ("Evicted", "CRITICAL", "phase=Failed, reason=Evicted", "Set ephemeral-storage limits"),
        ("Pending (Unschedulable)", "WARNING", "Insufficient resources", "Reduce requests or scale nodes"),
        ("ReadinessProbe Failure", "WARNING", "Events: Unhealthy + readiness", "Adjust probe thresholds"),
        ("LivenessProbe Failure", "CRITICAL", "Events: Unhealthy + liveness", "Increase initialDelaySeconds"),
        ("Init Container Failure", "WARNING", "Init in CrashLoopBackOff", "Fix init script / dependencies"),
        ("Volume Mount Error", "WARNING", "PVC pending / mount fail", "Fix PVC, StorageClass, zones"),
        ("Node NotReady Impact", "CRITICAL", "Pod on NotReady node", "Migrate pods, drain node"),
        ("Resource Quota Exceeded", "WARNING", "Rejected by admission", "Increase quota or reduce requests"),
    ]
    pattern_rows = ""
    for name, sev, detect, fix in pattern_data:
        sev_cls = "sev-crit" if sev == "CRITICAL" else "sev-warn"
        pattern_rows += f'''
        <tr>
          <td class="pname">{name}</td>
          <td class="{sev_cls}">{sev}</td>
          <td>{detect}</td>
          <td class="fix">{fix}</td>
        </tr>'''

    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>TCS Agentic AI — Use Case 2: Incident Response</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{ font-family: 'Inter', -apple-system, sans-serif; background: #f0f4f8; color: #1e293b; }}
  .page {{ max-width: 1400px; margin: 0 auto; padding: 24px; }}

  .header {{ background: linear-gradient(135deg, #1A1A2E 0%, #16213E 40%, #0F3460 70%, #E94560 100%);
    color: #fff; padding: 32px 40px; border-radius: 16px; margin-bottom: 24px; }}
  .header h1 {{ font-size: 24px; font-weight: 800; letter-spacing: -0.5px; }}
  .header h2 {{ font-size: 16px; font-weight: 500; color: rgba(255,255,255,0.8); margin-top: 4px; }}
  .header p {{ font-size: 13px; color: rgba(255,255,255,0.6); margin-top: 8px; line-height: 1.6; }}
  .header-metrics {{ display: flex; gap: 16px; margin-top: 16px; }}
  .hm {{ background: rgba(255,255,255,0.12); border: 1px solid rgba(255,255,255,0.2);
    padding: 10px 20px; border-radius: 10px; text-align: center; }}
  .hm .big {{ font-size: 22px; font-weight: 800; }}
  .hm .small {{ font-size: 10px; color: rgba(255,255,255,0.7); }}

  .section {{ margin-bottom: 28px; }}
  .section-title {{ font-size: 18px; font-weight: 700; color: #1A1A2E; margin-bottom: 14px;
    padding-bottom: 8px; border-bottom: 3px solid #E94560; display: inline-block; }}

  /* Workflow steps */
  .wf-grid {{ display: grid; grid-template-columns: repeat(5, 1fr); gap: 12px; }}
  .wf-step {{ background: #fff; border-radius: 10px; padding: 14px; position: relative;
    border: 1px solid #e2e8f0; box-shadow: 0 1px 3px rgba(0,0,0,0.06);
    display: flex; flex-direction: column; gap: 6px; }}
  .wf-step:nth-child(-n+5) {{ border-top: 3px solid #E94560; }}
  .wf-step:nth-child(n+6) {{ border-top: 3px solid #0F3460; }}
  .wf-num {{ font-size: 11px; font-weight: 800; color: #E94560;
    background: #FEE2E2; padding: 2px 8px; border-radius: 4px; display: inline-block; width: fit-content; }}
  .wf-step:nth-child(n+6) .wf-num {{ color: #0F3460; background: #DBEAFE; }}
  .wf-name {{ font-size: 12px; font-weight: 700; color: #1A1A2E; }}
  .wf-desc {{ font-size: 10px; color: #64748b; line-height: 1.4; }}
  .wf-arrow {{ display: flex; justify-content: center; align-items: center; margin: 8px 0;
    font-size: 20px; color: #94a3b8; }}

  /* Tables */
  table {{ width: 100%; border-collapse: collapse; background: #fff;
    border-radius: 12px; overflow: hidden; box-shadow: 0 1px 4px rgba(0,0,0,0.08); }}
  thead th {{ background: #1A1A2E; color: #fff; padding: 12px 14px;
    font-size: 11px; font-weight: 700; text-align: center; text-transform: uppercase;
    letter-spacing: 0.5px; border-right: 1px solid rgba(255,255,255,0.15); }}
  thead th:last-child {{ border-right: none; }}
  tbody td {{ padding: 10px 14px; border-bottom: 1px solid #e5e7eb;
    vertical-align: top; font-size: 12px; line-height: 1.5; }}
  tbody tr:nth-child(even) {{ background: #f8fafc; }}
  tbody tr:hover {{ background: #eff6ff; }}

  .phase {{ font-weight: 700; color: #1A1A2E; }}
  .manual {{ color: #DC2626; font-weight: 700; text-align: center; }}
  .auto {{ color: #16A34A; font-weight: 700; text-align: center; }}
  .saved {{ color: #16A34A; font-weight: 800; font-size: 14px; text-align: center; }}
  .pname {{ font-weight: 700; color: #1A1A2E; white-space: nowrap; }}
  .fix {{ color: #16A34A; font-weight: 600; }}
  .sev-crit {{ color: #DC2626; font-weight: 800; text-align: center; background: #FEE2E2 !important; }}
  .sev-warn {{ color: #E67E22; font-weight: 700; text-align: center; background: #FEF3C7 !important; }}

  .total-row {{ background: #1A1A2E !important; }}
  .total-row td {{ color: #fff; font-weight: 700; font-size: 13px; text-align: center; }}

  .footer {{ text-align: center; margin-top: 24px; padding: 16px;
    color: #94a3b8; font-size: 11px; }}

  @media print {{
    body {{ background: #fff; }}
    .page {{ padding: 0; max-width: 100%; }}
    .header {{ border-radius: 0; }}
    table {{ box-shadow: none; border: 1px solid #ccc; }}
    .wf-grid {{ grid-template-columns: repeat(5, 1fr); }}
  }}
</style>
</head>
<body>
<div class="page">

  <div class="header">
    <h1>TCS Agentic AI — Use Case 2</h1>
    <h2>End-to-End Incident Response: "Why is my pod crashing?"</h2>
    <p>Natural language question triggers a fully autonomous 10-stage pipeline: AI detects the error pattern,
       diagnoses root cause from 8 parallel K8s API calls, auto-creates a ServiceNow incident, proposes a
       targeted fix with dry-run validation, applies the fix, and verifies recovery — all in under 5 minutes.</p>
    <div class="header-metrics">
      <div class="hm"><div class="big">10</div><div class="small">Pipeline Stages</div></div>
      <div class="hm"><div class="big">12</div><div class="small">Error Patterns</div></div>
      <div class="hm"><div class="big">8</div><div class="small">Parallel API Calls</div></div>
      <div class="hm"><div class="big">&lt; 5 min</div><div class="small">End-to-End MTTR</div></div>
      <div class="hm"><div class="big">97%</div><div class="small">Time Saved vs Manual</div></div>
    </div>
  </div>

  <div class="section">
    <div class="section-title">10-Stage Workflow: What Happens When You Ask</div>
    <div class="wf-grid">{lifecycle_html}</div>
  </div>

  <div class="section">
    <div class="section-title">Manual SRE vs. TCS Agentic AI</div>
    <table>
      <thead>
        <tr><th>Phase</th><th>Manual (SRE Time)</th><th>With TCS Agentic AI</th><th>Saved</th></tr>
      </thead>
      <tbody>
        {comp_rows}
        <tr class="total-row">
          <td>TOTAL</td>
          <td style="color:#FEE2E2;">2.5–5.5 hours</td>
          <td style="color:#DCFCE7;">&lt; 5 minutes</td>
          <td style="color:#DCFCE7;">~97%</td>
        </tr>
      </tbody>
    </table>
  </div>

  <div class="section">
    <div class="section-title">12 Error Patterns Detected by Pod Doctor</div>
    <table>
      <thead>
        <tr><th>Pattern</th><th>Severity</th><th>How Detected</th><th>Auto-Fix</th></tr>
      </thead>
      <tbody>{pattern_rows}</tbody>
    </table>
  </div>

  <div class="footer">
    Powered by TCS &mdash; Tata Consultancy Services. All rights reserved. &nbsp;|&nbsp; Generated for hackathon demonstration purposes.
  </div>

</div>
</body>
</html>'''

    html_path = os.path.join(OUT_DIR, "TCS-Agentic-AI-Incident-Response-UseCase.html")
    with open(html_path, "w") as f:
        f.write(html)
    print(f"HTML saved: {html_path}")
    return html_path


# ═══════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    generate_html()
    print(f"\nDone! Files generated:")
    print(f"  Excel: {excel_path}")
